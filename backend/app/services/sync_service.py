"""
Сервис синхронизации данных с маркетплейсов
Загружает данные из WB и Ozon в Supabase
"""
import logging
from datetime import datetime, timedelta
from typing import Optional
from decimal import Decimal

from .wb_client import WildberriesClient
from .ozon_client import OzonClient, OzonPerformanceClient
from ..db.supabase import get_supabase_client
from ..config import get_settings
from ..plans import get_plan

logger = logging.getLogger(__name__)


class SyncService:
    """Сервис синхронизации данных с маркетплейсов в Supabase"""

    def __init__(self, user_id: str | None = None):
        settings = get_settings()
        self.supabase = get_supabase_client()
        if not user_id:
            raise ValueError("SyncService requires user_id — multi-tenant safety")
        self.user_id = user_id

        # Загружаем токены: сначала из БД (per-user), fallback на .env
        wb_tok, oz_cid, oz_key, oz_perf_cid, oz_perf_sec = self._load_tokens(user_id, settings)
        self.wb_client = WildberriesClient(wb_tok)
        self.ozon_client = OzonClient(oz_cid, oz_key)
        self.ozon_perf_client = OzonPerformanceClient(oz_perf_cid, oz_perf_sec)

        # Barcodes и Ozon SKU map загружаются из БД (mp_products)
        self._barcodes_cache: list[str] | None = None
        self._ozon_sku_map_cache: dict[str, str] | None = None

    def _load_tokens(self, user_id: str | None, settings) -> tuple:
        """Загрузить токены: БД (per-user) → fallback на .env."""
        from ..crypto import decrypt_token

        if user_id:
            try:
                result = (
                    self.supabase.table("mp_user_tokens")
                    .select("wb_api_token, ozon_client_id, ozon_api_key, ozon_perf_client_id, ozon_perf_secret")
                    .eq("user_id", user_id)
                    .limit(1)
                    .execute()
                )
                if result.data:
                    row = result.data[0]
                    return (
                        decrypt_token(row.get("wb_api_token") or "") or settings.wb_api_token,
                        decrypt_token(row.get("ozon_client_id") or "") or settings.ozon_client_id,
                        decrypt_token(row.get("ozon_api_key") or "") or settings.ozon_api_key,
                        decrypt_token(row.get("ozon_perf_client_id") or "") or settings.ozon_performance_client_id,
                        decrypt_token(row.get("ozon_perf_secret") or "") or settings.ozon_performance_client_secret,
                    )
            except Exception as e:
                logger.warning(f"Failed to load user tokens from DB, falling back to .env: {e}")

        # Fallback: .env
        return (
            settings.wb_api_token,
            settings.ozon_client_id,
            settings.ozon_api_key,
            settings.ozon_performance_client_id,
            settings.ozon_performance_client_secret,
        )

    @property
    def barcodes(self) -> list[str]:
        """Загрузить штрихкоды товаров из БД (кеш на время жизни SyncService)."""
        if self._barcodes_cache is None:
            products_map = self._get_products_map()
            self._barcodes_cache = [
                bc for bc in products_map.keys()
                if bc and bc != "WB_ACCOUNT"
            ]
        return self._barcodes_cache

    @property
    def ozon_sku_map(self) -> dict[str, str]:
        """Загрузить маппинг Ozon SKU → barcode из БД (поле ozon_sku в mp_products)."""
        if self._ozon_sku_map_cache is None:
            products_map = self._get_products_map()
            self._ozon_sku_map_cache = {}
            for barcode, product in products_map.items():
                ozon_sku = product.get("ozon_sku")
                if ozon_sku and barcode:
                    self._ozon_sku_map_cache[str(ozon_sku)] = barcode
        return self._ozon_sku_map_cache

    def _log_sync(self, marketplace: str, sync_type: str, status: str,
                  records_count: int = 0, error_message: str = None,
                  started_at: datetime = None):
        """Записать лог синхронизации в БД"""
        row = {
            "marketplace": marketplace,
            "sync_type": sync_type,
            "status": status,
            "records_count": records_count,
            "error_message": error_message,
            "started_at": started_at.isoformat() if started_at else None,
            "finished_at": datetime.now().isoformat(),
        }
        if self.user_id:
            row["user_id"] = self.user_id
        self.supabase.table("mp_sync_log").insert(row).execute()

    def _get_product_id_by_barcode(self, barcode: str) -> Optional[str]:
        """Получить UUID товара по штрихкоду"""
        query = self.supabase.table("mp_products").select("id").eq("barcode", barcode).limit(1)
        if self.user_id:
            query = query.eq("user_id", self.user_id)
        result = query.execute()
        if result.data:
            return result.data[0]["id"]
        return None

    def _get_products_map(self) -> dict:
        """Получить словарь товаров {barcode: {id, wb_nm_id, ozon_product_id}}"""
        query = self.supabase.table("mp_products").select("*").limit(500)
        if self.user_id:
            query = query.eq("user_id", self.user_id)
        result = query.execute()
        return {p["barcode"]: p for p in result.data}

    def _get_or_create_system_product_id(self, barcode: str, name: str) -> str:
        """
        Создаёт (если нужно) системный товар-заглушку для операций WB без привязки к товару
        (например, хранение/прочие удержания на уровне аккаунта).

        Важно: purchase_price в схеме NOT NULL → ставим 0.
        """
        query = self.supabase.table("mp_products").select("id").eq("barcode", barcode).limit(1)
        if self.user_id:
            query = query.eq("user_id", self.user_id)
        existing = query.execute()
        if existing.data:
            return existing.data[0]["id"]

        row = {
            "barcode": barcode,
            "name": name,
            "purchase_price": 0,
            "updated_at": datetime.now().isoformat(),
        }
        if self.user_id:
            row["user_id"] = self.user_id
        inserted = self.supabase.table("mp_products").insert(row).execute()
        # Supabase returns inserted row(s) in data for insert
        if inserted.data:
            return inserted.data[0]["id"]

        # Fallback: try re-select (in case insert returned empty but succeeded)
        query = self.supabase.table("mp_products").select("id").eq("barcode", barcode).limit(1)
        if self.user_id:
            query = query.eq("user_id", self.user_id)
        existing = query.execute()
        if existing.data:
            return existing.data[0]["id"]
        raise RuntimeError(f"Failed to create system product for barcode={barcode}")

    # ==================== HELPERS: PLAN & PAGINATION ====================

    def _get_user_plan(self) -> dict:
        """Get user's subscription plan config."""
        plan_name = "free"
        if self.user_id:
            try:
                result = (
                    self.supabase.table("mp_user_subscriptions")
                    .select("plan")
                    .eq("user_id", self.user_id)
                    .limit(1)
                    .execute()
                )
                if result.data:
                    plan_name = result.data[0]["plan"]
            except Exception as e:
                logger.warning(f"Failed to load user plan, defaulting to free: {e}")
        return get_plan(plan_name)

    def _get_max_sku(self) -> int | None:
        """Get max SKU count from user's plan (None = unlimited)."""
        return self._get_user_plan().get("max_sku")

    def _get_allowed_marketplaces(self) -> list[str]:
        """Get allowed marketplaces from user's plan."""
        return self._get_user_plan().get("marketplaces", ["wb"])

    async def _fetch_all_wb_cards(self) -> list[dict]:
        """Paginated fetch of ALL WB cards using cursor-based pagination."""
        all_cards = []
        cursor = {"limit": 100}

        while True:
            result = await self.wb_client.get_cards_list(cursor=cursor)
            if not result:
                break
            cards = result.get("cards", [])
            if not cards:
                break
            all_cards.extend(cards)

            # Cursor for next page: {total, updatedAt, nmID}
            next_cursor = result.get("cursor", {})
            total = next_cursor.get("total", 0)
            if len(all_cards) >= total or not next_cursor.get("nmID"):
                break
            cursor = {
                "limit": 100,
                "updatedAt": next_cursor.get("updatedAt"),
                "nmID": next_cursor.get("nmID"),
            }

        logger.info(f"WB: fetched {len(all_cards)} cards total")
        return all_cards

    async def _fetch_all_ozon_products(self) -> list[dict]:
        """Paginated fetch of ALL Ozon products using last_id pagination."""
        all_products = []
        last_id = ""

        while True:
            result = await self.ozon_client.get_product_list(limit=100, last_id=last_id)
            items = result.get("result", {}).get("items", [])
            if not items:
                break
            all_products.extend(items)
            last_id = result.get("result", {}).get("last_id", "")
            if not last_id:
                break

        logger.info(f"Ozon: fetched {len(all_products)} products total")
        return all_products

    # ==================== СИНХРОНИЗАЦИЯ ТОВАРОВ ====================

    async def sync_products(self) -> dict:
        """
        Синхронизация товаров с WB и Ozon — автообнаружение.
        Сканирует ВСЕ карточки/товары на маркетплейсах,
        обновляет существующие и создаёт новые mp_products строки
        (в рамках лимита SKU по тарифу).
        """
        started_at = datetime.now()
        updated_count = 0
        created_count = 0
        skipped_sku_limit = 0
        errors = []

        try:
            # a) Загружаем текущие товары из БД
            products_map = self._get_products_map()  # {barcode: {id, ...}}

            # b) Получаем план пользователя
            plan = self._get_user_plan()
            max_sku = plan.get("max_sku")  # None = unlimited
            allowed_marketplaces = plan.get("marketplaces", ["wb"])

            # Текущее кол-во SKU (без WB_ACCOUNT)
            current_sku_count = sum(
                1 for bc in products_map.keys()
                if bc and bc != "WB_ACCOUNT"
            )

            def _within_sku_limit() -> bool:
                nonlocal current_sku_count
                if max_sku is None:
                    return True
                return current_sku_count < max_sku

            # c) WB: auto-discover all cards
            if "wb" in allowed_marketplaces:
                try:
                    wb_cards = await self._fetch_all_wb_cards()

                    for card in wb_cards:
                        nm_id = card.get("nmID")
                        vendor_code = card.get("vendorCode")
                        card_name = card.get("title") or card.get("subjectName") or f"WB-{nm_id}"

                        for size in card.get("sizes", []):
                            for barcode in size.get("skus", []):
                                if not barcode:
                                    continue

                                if barcode in products_map:
                                    # UPDATE existing product
                                    query = self.supabase.table("mp_products").update({
                                        "wb_nm_id": nm_id,
                                        "wb_vendor_code": vendor_code,
                                        "updated_at": datetime.now().isoformat(),
                                    }).eq("barcode", barcode).eq("user_id", self.user_id)
                                    query.execute()
                                    updated_count += 1
                                    logger.info(f"WB: updated product {barcode} -> nm_id={nm_id}")
                                else:
                                    # INSERT new product (if within SKU limit)
                                    if not _within_sku_limit():
                                        skipped_sku_limit += 1
                                        logger.info(f"WB: skipped {barcode} — SKU limit reached ({max_sku})")
                                        continue

                                    row = {
                                        "barcode": barcode,
                                        "name": card_name,
                                        "purchase_price": 0,
                                        "wb_nm_id": nm_id,
                                        "wb_vendor_code": vendor_code,
                                        "updated_at": datetime.now().isoformat(),
                                    }
                                    if self.user_id:
                                        row["user_id"] = self.user_id
                                    self.supabase.table("mp_products").insert(row).execute()
                                    # Add to products_map so we don't double-insert
                                    products_map[barcode] = row
                                    current_sku_count += 1
                                    created_count += 1
                                    logger.info(f"WB: created product {barcode} ({card_name}) nm_id={nm_id}")

                except Exception as e:
                    error_msg = f"WB cards sync error: {e}"
                    logger.error(error_msg)
                    errors.append(error_msg)

            # d) Ozon: auto-discover all products
            ozon_product_ids = []
            if "ozon" in allowed_marketplaces:
                try:
                    ozon_products = await self._fetch_all_ozon_products()

                    # Collect offer_ids for new products that need names
                    new_ozon_offer_ids = []

                    for product in ozon_products:
                        product_id = product.get("product_id")
                        offer_id = product.get("offer_id")
                        if not offer_id:
                            continue

                        if offer_id in products_map:
                            # UPDATE existing product
                            query = self.supabase.table("mp_products").update({
                                "ozon_product_id": product_id,
                                "ozon_offer_id": offer_id,
                                "updated_at": datetime.now().isoformat(),
                            }).eq("barcode", offer_id).eq("user_id", self.user_id)
                            query.execute()
                            updated_count += 1
                            if product_id:
                                ozon_product_ids.append(product_id)
                            logger.info(f"Ozon: updated product {offer_id} -> product_id={product_id}")
                        else:
                            # INSERT new product (if within SKU limit)
                            if not _within_sku_limit():
                                skipped_sku_limit += 1
                                logger.info(f"Ozon: skipped {offer_id} — SKU limit reached ({max_sku})")
                                continue

                            new_ozon_offer_ids.append(offer_id)
                            row = {
                                "barcode": offer_id,
                                "name": f"Ozon-{offer_id}",  # placeholder, will update below
                                "purchase_price": 0,
                                "ozon_product_id": product_id,
                                "ozon_offer_id": offer_id,
                                "updated_at": datetime.now().isoformat(),
                            }
                            if self.user_id:
                                row["user_id"] = self.user_id
                            self.supabase.table("mp_products").insert(row).execute()
                            products_map[offer_id] = row
                            current_sku_count += 1
                            created_count += 1
                            if product_id:
                                ozon_product_ids.append(product_id)
                            logger.info(f"Ozon: created product {offer_id} product_id={product_id}")

                    # Get proper names for newly created Ozon products
                    if new_ozon_offer_ids:
                        try:
                            new_product_ids = [
                                products_map[oid].get("ozon_product_id")
                                for oid in new_ozon_offer_ids
                                if products_map.get(oid, {}).get("ozon_product_id")
                            ]
                            if new_product_ids:
                                info = await self.ozon_client.get_product_info(product_ids=new_product_ids)
                                name_items = info.get("items") or info.get("result", {}).get("items", [])
                                for item in name_items:
                                    item_offer_id = item.get("offer_id", "")
                                    item_name = item.get("name", "")
                                    if item_name and item_offer_id in new_ozon_offer_ids:
                                        query = self.supabase.table("mp_products").update({
                                            "name": item_name,
                                        }).eq("barcode", item_offer_id).eq("user_id", self.user_id)
                                        query.execute()
                                        logger.info(f"Ozon: updated name for {item_offer_id} -> {item_name}")
                        except Exception as e:
                            logger.warning(f"Failed to fetch Ozon product names for new products: {e}")

                except Exception as e:
                    error_msg = f"Ozon products sync error: {e}"
                    logger.error(error_msg)
                    errors.append(error_msg)

            # e) Fetch FBO SKU for Ozon products and save to ozon_sku
            sku_populated = set()  # offer_ids that got ozon_sku set

            if ozon_product_ids:
                # Method 1: try get_product_info (v3)
                try:
                    info = await self.ozon_client.get_product_info(product_ids=ozon_product_ids)
                    # v3 returns {"items": [...]}, not {"result": {"items": [...]}}
                    info_items = info.get("items") or info.get("result", {}).get("items", [])
                    for item in info_items:
                        item_offer_id = item.get("offer_id", "")
                        ozon_sku = None
                        # Try sources[].sku — prefer fbo, fallback to any source
                        for preferred_source in ("fbo", None):
                            for src in item.get("sources", []):
                                src_type = src.get("source", "")
                                if preferred_source and src_type != preferred_source:
                                    continue
                                if src.get("sku"):
                                    ozon_sku = str(src["sku"])
                                    break
                            if ozon_sku:
                                break
                        # Fallback: fbo_sku field
                        if not ozon_sku:
                            fbo_sku = item.get("fbo_sku")
                            if fbo_sku and int(fbo_sku) != 0:
                                ozon_sku = str(fbo_sku)
                        # Fallback: top-level sku
                        if not ozon_sku:
                            top_sku = item.get("sku")
                            if top_sku and int(top_sku) != 0:
                                ozon_sku = str(top_sku)

                        if ozon_sku and item_offer_id in products_map:
                            query = self.supabase.table("mp_products").update({
                                "ozon_sku": ozon_sku,
                            }).eq("barcode", item_offer_id).eq("user_id", self.user_id)
                            query.execute()
                            sku_populated.add(item_offer_id)
                            logger.info(f"Ozon: updated SKU via product_info {item_offer_id} -> ozon_sku={ozon_sku}")
                        elif item_offer_id:
                            logger.warning(f"Ozon: product_info returned no FBO SKU for {item_offer_id}")
                except Exception as e:
                    logger.warning(f"Failed to fetch Ozon product info for SKU mapping: {e}")

                # Method 2: fallback via warehouse stocks for products that still need ozon_sku
                offers_needing_sku = [
                    oid for oid in products_map
                    if oid != "WB_ACCOUNT"
                    and products_map[oid].get("ozon_product_id")
                    and oid not in sku_populated
                ]
                if offers_needing_sku:
                    logger.info(f"Ozon: {len(offers_needing_sku)} products still need FBO SKU, trying warehouse stocks fallback")
                    try:
                        wh_data = await self.ozon_client.get_all_stocks_on_warehouses()
                        rows = wh_data.get("result", {}).get("rows", [])
                        for row in rows:
                            item_code = row.get("item_code", "")
                            sku = row.get("sku")
                            if item_code and item_code in products_map and sku and item_code not in sku_populated:
                                ozon_sku = str(sku)
                                query = self.supabase.table("mp_products").update({
                                    "ozon_sku": ozon_sku,
                                }).eq("barcode", item_code).eq("user_id", self.user_id)
                                query.execute()
                                sku_populated.add(item_code)
                                logger.info(f"Ozon: updated SKU via warehouse fallback {item_code} -> ozon_sku={ozon_sku}")
                    except Exception as e:
                        logger.warning(f"Failed to fetch Ozon warehouse stocks for SKU mapping: {e}")

                if sku_populated:
                    logger.info(f"Ozon: populated ozon_sku for {len(sku_populated)} products")
                else:
                    logger.warning(f"Ozon: could not populate ozon_sku for any product!")

            # f) Try to fetch cost_price from Ozon API for products with purchase_price=0
            if "ozon" in allowed_marketplaces:
                zero_cc_offers = [
                    oid for oid, p in products_map.items()
                    if oid != "WB_ACCOUNT"
                    and p.get("ozon_product_id")
                    and (p.get("purchase_price") or 0) == 0
                ]
                if zero_cc_offers:
                    try:
                        prices_resp = await self.ozon_client.get_product_prices(
                            offer_ids=zero_cc_offers, limit=100
                        )
                        cc_updated = 0
                        for item in prices_resp.get("result", {}).get("items", []):
                            offer_id = item.get("offer_id", "")
                            # Check for cost_price in price object or top-level
                            price_obj = item.get("price", {}) if isinstance(item.get("price"), dict) else {}
                            cost_price = (
                                item.get("cost_price")
                                or item.get("purchase_price")
                                or price_obj.get("cost_price")
                                or price_obj.get("purchase_price")
                            )
                            if cost_price and offer_id in products_map:
                                try:
                                    cc_value = float(cost_price)
                                    if cc_value > 0:
                                        query = self.supabase.table("mp_products").update({
                                            "purchase_price": cc_value,
                                        }).eq("barcode", offer_id).eq("user_id", self.user_id)
                                        query.execute()
                                        cc_updated += 1
                                        logger.info(f"Ozon: set cost_price for {offer_id} -> {cc_value}")
                                except (ValueError, TypeError):
                                    pass
                        if cc_updated:
                            logger.info(f"Ozon: populated cost_price for {cc_updated} products from API")
                        else:
                            logger.info(f"Ozon: no cost_price found in prices API for {len(zero_cc_offers)} products")
                    except Exception as e:
                        logger.warning(f"Failed to fetch Ozon product prices for cost_price: {e}")

            # g) Invalidate caches after product sync
            self._barcodes_cache = None
            self._ozon_sku_map_cache = None

            # h) Return result
            total_processed = updated_count + created_count
            status = "success" if not errors else "partial"
            self._log_sync("all", "products", status, total_processed, started_at=started_at)
            result = {
                "status": status,
                "updated": updated_count,
                "created": created_count,
                "skipped_sku_limit": skipped_sku_limit,
            }
            if errors:
                result["errors"] = errors
            return result

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Ошибка синхронизации товаров: {error_msg}")
            self._log_sync("all", "products", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    # ==================== СИНХРОНИЗАЦИЯ ПРОДАЖ ====================

    async def sync_sales_wb(self, date_from: datetime, date_to: datetime = None) -> dict:
        """Синхронизация продаж с Wildberries"""
        started_at = datetime.now()
        date_to = date_to or datetime.now()
        records_count = 0

        try:
            products_map = self._get_products_map()

            # Источник истины для WB: финансовый отчёт reportDetailByPeriod (как в ЛК).
            # Важно: берём daily разрез, чтобы совпадать по датам.
            report = await self.wb_client.get_report_detail(date_from, date_to, period="daily")

            # Группируем по дате rr_dt, nm_id и fulfillment_type
            # sales/revenue считаем только по строкам "Продажа/Продажа".
            sales_agg = {}  # {(nm_id, date, ft): {sales, returns, revenue}}

            for row in report:
                nm_id = row.get("nm_id")
                date = (row.get("rr_dt") or "")[:10]
                if not nm_id or not date:
                    continue
                ft = self._determine_wb_fulfillment(row)
                key = (nm_id, date, ft)
                if key not in sales_agg:
                    sales_agg[key] = {"sales": 0, "returns": 0, "revenue": 0.0}

                doc = str(row.get("doc_type_name") or "")
                oper = str(row.get("supplier_oper_name") or "")

                try:
                    qty = int(row.get("quantity") or 0)
                except Exception:
                    qty = 0

                # Выкупы
                if doc == "Продажа" and oper == "Продажа":
                    if qty <= 0:
                        qty = 1
                    sales_agg[key]["sales"] += qty
                    sales_agg[key]["revenue"] += float(row.get("retail_amount", 0) or 0)
                # Возвраты/сторно (best-effort)
                elif "Возврат" in doc or "Возврат" in oper or "Сторно" in doc or "Сторно" in oper:
                    if qty <= 0:
                        qty = 1
                    sales_agg[key]["returns"] += abs(qty)

            # Сохраняем в БД (batch upsert по 500)
            sales_batch = []
            for (nm_id, date, ft), data in sales_agg.items():
                # Находим товар по nm_id
                product_id = None
                for barcode, product in products_map.items():
                    if product.get("wb_nm_id") == nm_id:
                        product_id = product["id"]
                        break

                if not product_id:
                    continue

                # В reportDetailByPeriod нет заказов как сущности.
                # Для совместимости считаем orders = sales + returns (как "продажи и возвраты").
                orders_count = int(data["sales"] + data["returns"])
                buyout_percent = round(data["sales"] / orders_count * 100, 2) if orders_count > 0 else None

                row = {
                    "product_id": product_id,
                    "marketplace": "wb",
                    "date": date,
                    "orders_count": orders_count,
                    "sales_count": data["sales"],
                    "returns_count": data["returns"],
                    "revenue": data["revenue"],
                    "buyout_percent": buyout_percent,
                    "fulfillment_type": ft,
                }
                if self.user_id:
                    row["user_id"] = self.user_id
                sales_batch.append(row)

            for i in range(0, len(sales_batch), 500):
                chunk = sales_batch[i:i+500]
                self.supabase.table("mp_sales").upsert(
                    chunk, on_conflict="user_id,product_id,marketplace,date,fulfillment_type"
                ).execute()
            records_count = len(sales_batch)

            # Получаем воронку (cart_adds) если есть nm_ids
            nm_ids = [p.get("wb_nm_id") for p in products_map.values() if p.get("wb_nm_id")]
            if nm_ids:
                try:
                    funnel = await self.wb_client.get_funnel(date_from, date_to, nm_ids)
                    for card in funnel.get("data", {}).get("cards", []):
                        nm_id = card.get("nmID")
                        # Находим product_id
                        product_id = None
                        for barcode, product in products_map.items():
                            if product.get("wb_nm_id") == nm_id:
                                product_id = product["id"]
                                break

                        if product_id:
                            for stat in card.get("statistics", {}).get("selectedPeriod", {}).get("conversions", []):
                                # Обновляем cart_adds за период
                                cart_adds = stat.get("addToCart", 0)
                                query = self.supabase.table("mp_sales").update({
                                    "cart_adds": cart_adds
                                }).eq("product_id", product_id).eq("marketplace", "wb").gte("date", date_from.strftime("%Y-%m-%d")).eq("user_id", self.user_id)
                                query.execute()
                except Exception as e:
                    logger.warning(f"Не удалось получить воронку WB: {e}")

            self._log_sync("wb", "sales", "success", records_count, started_at=started_at)
            return {"status": "success", "records": records_count}

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Ошибка синхронизации продаж WB: {error_msg}")
            self._log_sync("wb", "sales", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    async def sync_sales_ozon(self, date_from: datetime, date_to: datetime = None) -> dict:
        """Синхронизация продаж с Ozon (по дням)"""
        started_at = datetime.now()
        date_to = date_to or datetime.now()
        records_count = 0

        try:
            products_map = self._get_products_map()

            # Получаем аналитику с разбивкой по SKU и дням (с пагинацией)
            all_rows = []
            offset = 0
            page_limit = 1000
            while True:
                analytics = await self.ozon_client.get_analytics_data(
                    date_from, date_to,
                    dimensions=["sku", "day"],
                    metrics=["ordered_units", "revenue", "returns", "session_view", "hits_tocart", "delivered_units"],
                    limit=page_limit,
                    offset=offset
                )
                page_data = analytics.get("result", {}).get("data", [])
                all_rows.extend(page_data)
                if len(page_data) < page_limit:
                    break
                offset += page_limit
                if offset > 10000:  # safety limit
                    logger.warning(f"sync_sales_ozon: достигнут лимит пагинации offset={offset}")
                    break

            sales_batch = []
            for row in all_rows:
                dimensions = row.get("dimensions", [])
                metrics = row.get("metrics", [])

                if len(dimensions) < 2:
                    continue

                sku = str(dimensions[0].get("id"))
                day_str = dimensions[1].get("id", "")  # формат "2026-01-16"

                if not day_str:
                    continue

                # Находим barcode через маппинг SKU
                barcode = self.ozon_sku_map.get(sku)
                if not barcode:
                    logger.warning(f"Товар с Ozon SKU {sku} не найден в маппинге")
                    continue

                # Находим product_id по barcode
                product_id = None
                if barcode in products_map:
                    product_id = products_map[barcode]["id"]

                if not product_id:
                    logger.warning(f"Товар с barcode {barcode} не найден в БД")
                    continue

                # Метрики: [ordered_units, revenue, returns, session_view, hits_tocart, delivered_units]
                orders = int(metrics[0]) if len(metrics) > 0 else 0
                revenue = float(metrics[1]) if len(metrics) > 1 else 0
                returns = int(metrics[2]) if len(metrics) > 2 else 0
                views = int(metrics[3]) if len(metrics) > 3 else 0
                cart_adds = int(metrics[4]) if len(metrics) > 4 else 0
                delivered = int(metrics[5]) if len(metrics) > 5 else 0

                # Пропускаем дни без активности
                if orders == 0 and revenue == 0 and returns == 0:
                    continue

                # Используем delivered_units если доступен (точнее чем orders - returns)
                sales_count = delivered if delivered > 0 else max(0, orders - returns)

                # buyout% на основе лучших доступных данных
                if delivered > 0 and orders > 0:
                    buyout = round(delivered / orders * 100, 2)
                elif orders > 0:
                    buyout = round((orders - returns) / orders * 100, 2)
                else:
                    buyout = None

                upsert_row = {
                    "product_id": product_id,
                    "marketplace": "ozon",
                    "date": day_str,
                    "orders_count": orders,
                    "sales_count": sales_count,
                    "returns_count": returns,
                    "revenue": revenue,
                    "cart_adds": cart_adds,
                    "views": views,
                    "buyout_percent": buyout,
                    "fulfillment_type": "FBO",
                }
                if self.user_id:
                    upsert_row["user_id"] = self.user_id
                sales_batch.append(upsert_row)

            for i in range(0, len(sales_batch), 500):
                chunk = sales_batch[i:i+500]
                self.supabase.table("mp_sales").upsert(
                    chunk, on_conflict="user_id,product_id,marketplace,date,fulfillment_type"
                ).execute()
            records_count = len(sales_batch)

            self._log_sync("ozon", "sales", "success", records_count, started_at=started_at)
            return {"status": "success", "records": records_count}

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Ошибка синхронизации продаж Ozon: {error_msg}")
            self._log_sync("ozon", "sales", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    # ==================== СИНХРОНИЗАЦИЯ ОСТАТКОВ ====================

    async def sync_stocks_wb(self) -> dict:
        """Синхронизация остатков с Wildberries"""
        started_at = datetime.now()
        records_count = 0

        try:
            products_map = self._get_products_map()

            # WB stocks — срез по складам. Для "активного" остатка используем поле quantity.
            #
            # ВАЖНО: /api/v1/supplier/stocks параметр dateFrom ведёт себя как "изменения с даты",
            # поэтому ответ может быть НЕ полным снапшотом по всем складам/товарам.
            # Следствие: нельзя удалять отсутствующие в ответе склады, иначе будем терять остатки
            # на складах, где не было изменений (пример: "Электросталь" на скрине).
            #
            # Чтобы получить ПОЛНЫЙ срез "как в ЛК", WB рекомендует передавать максимально раннюю дату
            # (иначе при свежем dateFrom API возвращает только изменения по складам/товарам).
            stocks = await self.wb_client.get_stocks(datetime(2019, 6, 20))

            if not stocks:
                logger.warning("WB stocks: API returned empty list (skip DB update to avoid wiping data)")
                self._log_sync("wb", "stocks", "success", 0, started_at=started_at)
                return {"status": "success", "records": 0}

            barcode_to_pid: dict[str, str] = {}
            nm_to_pid: dict[int, str] = {}
            for bc, p in products_map.items():
                pid = p.get("id")
                if pid and bc:
                    barcode_to_pid[str(bc).strip()] = pid
                nm = p.get("wb_nm_id")
                if pid and nm:
                    try:
                        nm_int = int(nm)
                        # Если внезапно несколько product_id на один nm_id — логируем, но не валим.
                        if nm_int in nm_to_pid and nm_to_pid[nm_int] != pid:
                            logger.warning(f"WB stocks: multiple product_ids for nm_id={nm_int}: {nm_to_pid[nm_int]} vs {pid}")
                        else:
                            nm_to_pid[nm_int] = pid
                    except Exception:
                        pass

            def _pid_for_stock_row(row: dict) -> str | None:
                bc = str(row.get("barcode") or "").strip()
                if bc:
                    pid = barcode_to_pid.get(bc)
                    if pid:
                        return pid
                nm = row.get("nmId")
                if nm is not None:
                    try:
                        nm_int = int(nm)
                        return nm_to_pid.get(nm_int)
                    except Exception:
                        return None
                return None

            # Суммируем по (product_id, warehouse), т.к. WB может отдавать несколько строк на один товар
            # (разные barcodes/sizes) — нельзя "последней строкой" затирать значение.
            agg: dict[tuple[str, str], int] = {}
            for stock in stocks:
                pid = _pid_for_stock_row(stock)
                if not pid:
                    # диагностически полезно: есть строки WB, которые мы не смогли сопоставить
                    logger.info(f"WB stocks: unmapped row nmId={stock.get('nmId')} barcode={stock.get('barcode')}")
                    continue

                wh = str(stock.get("warehouseName") or "Unknown").strip() or "Unknown"
                try:
                    qty = int(stock.get("quantity") or 0)
                except Exception:
                    qty = 0
                if qty < 0:
                    qty = 0

                key = (pid, wh)
                agg[key] = agg.get(key, 0) + qty

            now_iso = datetime.now().isoformat()
            stocks_batch = []
            for (pid, wh), qty in agg.items():
                upsert_row = {
                    "product_id": pid,
                    "marketplace": "wb",
                    "warehouse": wh,
                    "quantity": qty,
                    "updated_at": now_iso,
                    "fulfillment_type": "FBO",
                }
                if self.user_id:
                    upsert_row["user_id"] = self.user_id
                stocks_batch.append(upsert_row)

            for i in range(0, len(stocks_batch), 500):
                chunk = stocks_batch[i:i+500]
                self.supabase.table("mp_stocks").upsert(
                    chunk,
                    on_conflict="user_id,product_id,marketplace,warehouse,fulfillment_type",
                ).execute()
            records_count = len(stocks_batch)

            self._log_sync("wb", "stocks", "success", records_count, started_at=started_at)
            self._save_stock_snapshot("wb")
            return {"status": "success", "records": records_count}

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Ошибка синхронизации остатков WB: {error_msg}")
            self._log_sync("wb", "stocks", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    def _save_stock_snapshot(self, marketplace: str) -> None:
        """Save daily snapshot of current stock quantities per product for history chart."""
        if not self.user_id:
            return
        try:
            today = datetime.now().strftime("%Y-%m-%d")
            # Get current stocks for this marketplace (with fulfillment_type)
            result = self.supabase.table("mp_stocks") \
                .select("product_id, quantity, fulfillment_type") \
                .eq("user_id", self.user_id) \
                .eq("marketplace", marketplace) \
                .limit(5000) \
                .execute()

            # Aggregate by (product_id, fulfillment_type) — sum across warehouses
            totals: dict[tuple[str, str], int] = {}
            for row in result.data:
                pid = row.get("product_id")
                ft = row.get("fulfillment_type", "FBO")
                if pid:
                    key = (pid, ft)
                    totals[key] = totals.get(key, 0) + (row.get("quantity") or 0)

            # Upsert daily snapshots (batch по 500)
            snapshot_batch = []
            for (pid, ft), qty in totals.items():
                snapshot_batch.append({
                    "user_id": self.user_id,
                    "product_id": pid,
                    "marketplace": marketplace,
                    "date": today,
                    "total_quantity": qty,
                    "fulfillment_type": ft,
                })

            for i in range(0, len(snapshot_batch), 500):
                chunk = snapshot_batch[i:i+500]
                self.supabase.table("mp_stock_snapshots").upsert(
                    chunk,
                    on_conflict="user_id,product_id,marketplace,date,fulfillment_type",
                ).execute()

            logger.info(f"Stock snapshot saved: {marketplace}, {len(snapshot_batch)} product×ft combos")
        except Exception as e:
            # Don't fail the sync if snapshot fails
            logger.warning(f"Failed to save stock snapshot ({marketplace}): {e}")

    async def diagnose_stocks_wb(self, days_back: int = 365) -> dict:
        """
        Диагностика остатков WB: сравнение WB API vs mp_stocks.

        Возвращает:
        - сколько строк пришло из WB API, сколько не сматчилось
        - totals по баркодам
        - расхождения по складам (missing/extra) между WB и БД
        """
        products_map = self._get_products_map()

        barcode_to_pid: dict[str, str] = {}
        pid_to_barcode: dict[str, str] = {}
        pid_to_name: dict[str, str] = {}
        nm_to_pid: dict[int, str] = {}
        for bc, p in products_map.items():
            pid = p.get("id")
            if pid and bc:
                bcs = str(bc).strip()
                barcode_to_pid[bcs] = pid
                pid_to_barcode[pid] = bcs
            if pid:
                pid_to_name[pid] = str(p.get("name") or "")
            nm = p.get("wb_nm_id")
            if pid and nm:
                try:
                    nm_to_pid[int(nm)] = pid
                except Exception:
                    pass

        # Для точного мэтчинга с ЛК WB нужно запрашивать с максимально ранней даты.
        # days_back оставляем как override для экспериментов/отладки.
        if days_back >= 365:
            wb_rows = await self.wb_client.get_stocks(datetime(2019, 6, 20))
        else:
            wb_rows = await self.wb_client.get_stocks(datetime.now() - timedelta(days=days_back))
        if not wb_rows:
            wb_rows = []

        def _pid_for_row(row: dict) -> str | None:
            bc = str(row.get("barcode") or "").strip()
            if bc:
                pid = barcode_to_pid.get(bc)
                if pid:
                    return pid
            nm = row.get("nmId")
            if nm is not None:
                try:
                    return nm_to_pid.get(int(nm))
                except Exception:
                    return None
            return None

        # WB agg
        wb_unmapped = 0
        wb_by_pid_wh: dict[tuple[str, str], int] = {}
        wb_by_pid_total: dict[str, int] = {}
        for row in wb_rows:
            pid = _pid_for_row(row)
            if not pid:
                wb_unmapped += 1
                continue
            wh = str(row.get("warehouseName") or "Unknown").strip() or "Unknown"
            try:
                qty = int(row.get("quantity") or 0)
            except Exception:
                qty = 0
            if qty < 0:
                qty = 0
            wb_by_pid_wh[(pid, wh)] = wb_by_pid_wh.get((pid, wh), 0) + qty
            wb_by_pid_total[pid] = wb_by_pid_total.get(pid, 0) + qty

        # DB agg
        db_query = (
            self.supabase.table("mp_stocks")
            .select("product_id, warehouse, quantity, updated_at, mp_products(name, barcode)")
            .eq("marketplace", "wb")
            .limit(5000)
        )
        if self.user_id:
            db_query = db_query.eq("user_id", self.user_id)
        db_rows = db_query.execute()
        db_by_pid_wh: dict[tuple[str, str], int] = {}
        db_by_pid_total: dict[str, int] = {}
        last_updated_by_pid: dict[str, str] = {}
        for row in (db_rows.data or []):
            pid = row.get("product_id")
            if not pid:
                continue
            wh = str(row.get("warehouse") or "Unknown").strip() or "Unknown"
            try:
                qty = int(row.get("quantity") or 0)
            except Exception:
                qty = 0
            if qty < 0:
                qty = 0
            db_by_pid_wh[(pid, wh)] = db_by_pid_wh.get((pid, wh), 0) + qty
            db_by_pid_total[pid] = db_by_pid_total.get(pid, 0) + qty
            upd = row.get("updated_at")
            if upd:
                cur = last_updated_by_pid.get(pid)
                if not cur or str(upd) > str(cur):
                    last_updated_by_pid[pid] = str(upd)

        # diffs per product
        all_pids = sorted(set(list(wb_by_pid_total.keys()) + list(db_by_pid_total.keys())))
        diffs = []
        for pid in all_pids:
            wb_total = wb_by_pid_total.get(pid, 0)
            db_total = db_by_pid_total.get(pid, 0)
            delta = wb_total - db_total

            wb_whs = {wh for (p, wh) in wb_by_pid_wh.keys() if p == pid}
            db_whs = {wh for (p, wh) in db_by_pid_wh.keys() if p == pid}
            missing_in_db = sorted(wb_whs - db_whs)
            extra_in_db = sorted(db_whs - wb_whs)

            # only include interesting diffs
            if delta != 0 or missing_in_db or extra_in_db:
                diffs.append(
                    {
                        "product_id": pid,
                        "barcode": pid_to_barcode.get(pid),
                        "product_name": pid_to_name.get(pid),
                        "wb_total": wb_total,
                        "db_total": db_total,
                        "delta": delta,
                        "missing_warehouses_in_db": missing_in_db,
                        "extra_warehouses_in_db": extra_in_db,
                        "db_last_updated_at": last_updated_by_pid.get(pid),
                    }
                )

        diffs.sort(key=lambda x: abs(int(x.get("delta") or 0)), reverse=True)

        return {
            "status": "success",
            "fetched_at": datetime.now().isoformat(),
            "days_back": days_back,
            "wb": {
                "rows": len(wb_rows),
                "unmapped_rows": wb_unmapped,
                "products_seen": len(wb_by_pid_total),
            },
            "db": {
                "rows": len(db_by_pid_wh),
                "products_seen": len(db_by_pid_total),
            },
            "diffs": diffs,
        }

    async def sync_stocks_ozon(self) -> dict:
        """Синхронизация остатков с Ozon"""
        started_at = datetime.now()
        records_count = 0

        try:
            products_map = self._get_products_map()

            # Ozon stocks API часто отдаёт items=[] при filter={}, даже если остатки есть.
            # Поэтому пробуем несколько стратегий:
            # 1) точечно по нашим offer_id (у нас offer_id = barcode) + visibility=ALL
            # 2) точечно по offer_id без visibility
            # 3) общая выборка visibility=ALL
            # Для stocks-запросов используем идентификаторы из Ozon API (sync_products),
            # а не предполагаем, что offer_id == barcode.
            offer_ids: list[str] = []
            product_ids: list[int] = []
            for bc, p in products_map.items():
                if not bc or bc == "WB_ACCOUNT":
                    continue
                oid = (p.get("ozon_offer_id") or "").strip()
                if oid:
                    offer_ids.append(oid)
                pid = p.get("ozon_product_id")
                if pid:
                    try:
                        product_ids.append(int(pid))
                    except Exception:
                        pass

            tried: list[tuple[str, dict]] = []

            async def _fetch_items(flt: dict) -> list[dict]:
                tried.append(("v4/product/info/stocks", flt))
                res = await self.ozon_client.get_all_stocks(filter=flt)
                return (res or {}).get("result", {}).get("items", []) or []

            # 1) product_id (самый надёжный для FBO/"Склад Ozon")
            stocks: list[dict] = []
            if product_ids:
                stocks = await _fetch_items({"visibility": "ALL", "product_id": product_ids})
                if not stocks:
                    stocks = await _fetch_items({"product_id": product_ids})

            # 2) offer_id
            if not stocks:
                stocks = await _fetch_items({"visibility": "ALL", "offer_id": offer_ids})
            if not stocks:
                stocks = await _fetch_items({"offer_id": offer_ids})
            if not stocks:
                stocks = await _fetch_items({"visibility": "ALL"})
            # Fallback: если по offer_id пусто, пробуем через product_id из /v3/product/info/list
            # (в некоторых аккаунтах/режимах Ozon stocks лучше работает по product_id, особенно для "Склад Ozon" (FBO)).
            if not stocks and offer_ids and not product_ids:
                try:
                    info = await self.ozon_client.get_product_info(offer_ids=offer_ids)
                    items = (info.get("items") or info.get("result", {}).get("items", [])) if isinstance(info, dict) else []
                    product_ids_fb = [it.get("product_id") for it in items if it.get("product_id")]
                    if product_ids_fb:
                        stocks = await _fetch_items({"visibility": "ALL", "product_id": product_ids_fb})
                        if not stocks:
                            stocks = await _fetch_items({"product_id": product_ids_fb})
                except Exception as _e:
                    # Не валим sync полностью из-за fallback — просто продолжаем с пустыми stocks.
                    logger.warning(f"Ozon stocks fallback by product_id failed: {_e}")

            # === FBO fallback: Analytics /v2/analytics/stock_on_warehouses ===
            # В некоторых аккаунтах /v4/product/info/stocks отдаёт только "Мои склады" (FBS) или пусто,
            # тогда как FBO ("Склад Ozon") доступен через analytics отчёт.
            if not stocks:
                tried.append(("v2/analytics/stock_on_warehouses", {"warehouse_type": "ALL"}))
                try:
                    report = await self.ozon_client.get_all_stocks_on_warehouses(warehouse_type="ALL", limit=100)
                    rows = (report or {}).get("result", {}).get("rows", []) or []

                    # Индексы для маппинга
                    offer_to_pid: dict[str, str] = {}
                    for _, p in products_map.items():
                        oid = str(p.get("ozon_offer_id") or "").strip()
                        if oid and p.get("id"):
                            offer_to_pid[oid] = p["id"]

                    now_iso = datetime.now().isoformat()
                    stocks_batch: list[dict] = []
                    for row in rows:
                        item_code = str(row.get("item_code") or "").strip()
                        sku = row.get("sku")
                        warehouse = str(row.get("warehouse_name") or "Ozon").strip() or "Ozon"

                        # Qty: близко к "Активный сток" в ЛК: free_to_sell_amount
                        try:
                            qty = int(row.get("free_to_sell_amount") or 0)
                        except Exception:
                            qty = 0

                        # product_id (UUID) маппим по offer_id (item_code) или по sku-map (временная схема проекта)
                        product_id: str | None = None
                        if item_code:
                            product_id = offer_to_pid.get(item_code) or products_map.get(item_code, {}).get("id")

                        if not product_id and sku is not None:
                            barcode = self.ozon_sku_map.get(str(sku))
                            if barcode:
                                product_id = products_map.get(barcode, {}).get("id")

                        if not product_id:
                            logger.warning(f"Ozon stock_on_warehouses: unmapped row sku={sku} item_code={item_code}")
                            continue

                        upsert_row = {
                            "product_id": product_id,
                            "marketplace": "ozon",
                            "warehouse": warehouse,
                            "quantity": qty,
                            "updated_at": now_iso,
                            "fulfillment_type": "FBO",
                        }
                        if self.user_id:
                            upsert_row["user_id"] = self.user_id
                        stocks_batch.append(upsert_row)

                    for i in range(0, len(stocks_batch), 500):
                        chunk = stocks_batch[i:i+500]
                        self.supabase.table("mp_stocks").upsert(
                            chunk,
                            on_conflict="user_id,product_id,marketplace,warehouse,fulfillment_type",
                        ).execute()
                    records_count = len(stocks_batch)

                    self._log_sync("ozon", "stocks", "success", records_count, started_at=started_at)
                    self._save_stock_snapshot("ozon")
                    return {"status": "success", "records": records_count, "source": "v2/analytics/stock_on_warehouses"}
                except Exception as e:
                    logger.warning(f"Ozon analytics stock_on_warehouses failed: {e}")

            now_iso = datetime.now().isoformat()
            stocks_batch: list[dict] = []
            for stock in stocks:
                product_id_ozon = stock.get("product_id")
                offer_id = stock.get("offer_id")

                # Находим товар
                product_id = None
                for barcode, product in products_map.items():
                    if product.get("ozon_product_id") == product_id_ozon or product.get("ozon_offer_id") == offer_id:
                        product_id = product["id"]
                        break

                if not product_id:
                    continue

                # Суммируем остатки по складам
                for warehouse_stock in stock.get("stocks", []):
                    warehouse = warehouse_stock.get("warehouse_name", "FBS")
                    present = int(warehouse_stock.get("present", 0) or 0)
                    reserved = int(warehouse_stock.get("reserved", 0) or 0)
                    # Ближе к "доступно к продаже": present - reserved (если reserved есть).
                    quantity = max(present - reserved, 0)

                    # Определяем тип фулфилмента по типу склада
                    wh_type = warehouse_stock.get("type", "")
                    wh_name_lower = warehouse.lower()
                    if wh_type == "fbo" or "ozon" in wh_name_lower:
                        ft = "FBO"
                    else:
                        ft = "FBS"

                    upsert_row = {
                        "product_id": product_id,
                        "marketplace": "ozon",
                        "warehouse": warehouse,
                        "quantity": quantity,
                        "updated_at": now_iso,
                        "fulfillment_type": ft,
                    }
                    if self.user_id:
                        upsert_row["user_id"] = self.user_id
                    stocks_batch.append(upsert_row)

            for i in range(0, len(stocks_batch), 500):
                chunk = stocks_batch[i:i+500]
                self.supabase.table("mp_stocks").upsert(
                    chunk,
                    on_conflict="user_id,product_id,marketplace,warehouse,fulfillment_type",
                ).execute()
            records_count = len(stocks_batch)

            # Если по Ozon пусто — логируем как success с 0, но добавляем диагностическое сообщение
            # в стандартный error_message нельзя (это поле для error), поэтому пишем в лог сервера.
            if not stocks:
                logger.warning(f"Ozon stocks: items=0. Tried filters: {tried!r}")

            self._log_sync("ozon", "stocks", "success", records_count, started_at=started_at)
            self._save_stock_snapshot("ozon")
            return {"status": "success", "records": records_count}

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Ошибка синхронизации остатков Ozon: {error_msg}")
            self._log_sync("ozon", "stocks", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    # ==================== СИНХРОНИЗАЦИЯ УДЕРЖАНИЙ ====================

    async def sync_costs_wb(self, date_from: datetime, date_to: datetime = None) -> dict:
        """Синхронизация удержаний с Wildberries"""
        started_at = datetime.now()
        date_to = date_to or datetime.now()
        records_count = 0

        try:
            products_map = self._get_products_map()

            # Источник истины для WB: reportDetailByPeriod (как в ЛК).
            report = await self.wb_client.get_report_detail(date_from, date_to, period="daily")

            # Агрегация для mp_costs (положительные суммы расходов)
            costs_agg = {}  # {(product_id, date, ft): {commission, logistics, storage, ...}}

            # Гранулярные данные для mp_costs_details (tree-view)
            details_agg = {}  # {(product_id, date, ft, category, subcategory): amount}
            payout_by_key = {}  # {(product_id, date, ft): payout_sum}

            WB_SYSTEM_BARCODE = "WB_ACCOUNT"
            WB_SYSTEM_NAME = "WB: вне разреза товаров"
            wb_system_pid: str | None = None

            def _pid_by_nm(nm_id: int | None):
                if nm_id is None:
                    return None
                for _, product in products_map.items():
                    if product.get("wb_nm_id") == nm_id:
                        return product["id"]
                return None

            def _pid_for_row(row: dict) -> str:
                nonlocal wb_system_pid
                nm_id = row.get("nm_id")
                if nm_id:
                    pid = _pid_by_nm(nm_id)
                    if pid:
                        return pid

                # Fallback by barcode (если nm_id не замаплен, но barcode присутствует)
                bc = str(row.get("barcode") or "").strip()
                if bc:
                    pid = products_map.get(bc, {}).get("id")
                    if pid:
                        return pid

                # Account-level row (хранение/прочее) — складываем в системный "товар"
                if wb_system_pid is None:
                    wb_system_pid = self._get_or_create_system_product_id(WB_SYSTEM_BARCODE, WB_SYSTEM_NAME)
                    products_map[WB_SYSTEM_BARCODE] = {"id": wb_system_pid}
                return wb_system_pid

            for row in report:
                date = (row.get("rr_dt") or "")[:10]
                if not date:
                    continue
                product_id = _pid_for_row(row)
                ft = self._determine_wb_fulfillment(row)

                key = (product_id, date, ft)
                if key not in costs_agg:
                    costs_agg[key] = {
                        "commission": 0.0,
                        "logistics": 0.0,
                        "storage": 0.0,
                        "promotion": 0.0,
                        "penalties": 0.0,
                        "acquiring": 0.0,
                        "other": 0.0,
                    }
                    payout_by_key[key] = 0.0

                # === 1) Детализация для tree-view ===
                operation_type = str(row.get("doc_type_name") or row.get("supplier_oper_name") or "WB")
                operation_id = str(row.get("rrd_id") or "")
                for rec in self._classify_wb_report_row(row):
                    dkey = (product_id, date, ft, rec["category"], rec["subcategory"])
                    if dkey not in details_agg:
                        details_agg[dkey] = {"amount": 0.0, "operation_type": operation_type, "operation_id": operation_id}
                    details_agg[dkey]["amount"] += float(rec["amount"] or 0)

                # К перечислению (для балансировки дерева под 100% мэтч с ЛК WB)
                payout_by_key[key] += float(row.get("ppvz_for_pay", 0) or 0)

                # === 2) Агрегация для mp_costs (без знаков, только расходы) ===
                vw = abs(float(row.get("ppvz_vw", 0) or 0))
                vw_nds = abs(float(row.get("ppvz_vw_nds", 0) or 0))
                costs_agg[key]["commission"] += (vw + vw_nds)

                # Логистика как расход — только "Услуги по доставке товара покупателю".
                # Возмещения (ПВЗ, перевозка/складские операции) — это отдельные начисления в отчёте, их не считаем расходом.
                delivery = abs(float(row.get("delivery_rub", 0) or 0))
                costs_agg[key]["logistics"] += delivery

                costs_agg[key]["acquiring"] += abs(float(row.get("acquiring_fee", 0) or 0))
                costs_agg[key]["storage"] += abs(float(row.get("storage_fee", 0) or 0))
                costs_agg[key]["penalties"] += abs(float(row.get("penalty", 0) or 0))

                # Прочие удержания/операции (best-effort, чтобы сходилась общая сумма)
                costs_agg[key]["other"] += abs(float(row.get("deduction", 0) or 0))
                costs_agg[key]["other"] += abs(float(row.get("acceptance", 0) or 0))

            # Доводка до "К перечислению" через кабинетную сущность "Прочие удержания/выплаты".
            # Это НЕ "затычка": в ЛК WB есть отдельная колонка "Прочие удержания/выплаты".
            for (pid, date, ft) in payout_by_key.keys():
                payout = payout_by_key[(pid, date, ft)]
                details_sum = 0.0
                for (p, d, f, _, _), data in details_agg.items():
                    if p == pid and d == date and f == ft:
                        details_sum += float(data["amount"] or 0)
                diff = round(payout - details_sum, 2)
                if abs(diff) >= 0.01:
                    dkey = (pid, date, ft, "Прочие удержания/выплаты", "Прочие удержания/выплаты")
                    if dkey not in details_agg:
                        details_agg[dkey] = {"amount": 0.0, "operation_type": "WB_RESIDUAL", "operation_id": ""}
                    details_agg[dkey]["amount"] += diff

            # Сохраняем mp_costs (batch upsert по 500)
            costs_batch = []
            for (pid, date, ft), costs in costs_agg.items():
                upsert_row = {
                    "product_id": pid,
                    "marketplace": "wb",
                    "date": date,
                    "commission": round(costs["commission"], 2),
                    "logistics": round(costs["logistics"], 2),
                    "storage": round(costs["storage"], 2),
                    "promotion": round(costs["promotion"], 2),
                    "penalties": round(costs["penalties"], 2),
                    "acquiring": round(costs["acquiring"], 2),
                    "other_costs": round(costs["other"], 2),
                    "fulfillment_type": ft,
                }
                if self.user_id:
                    upsert_row["user_id"] = self.user_id
                costs_batch.append(upsert_row)

            for i in range(0, len(costs_batch), 500):
                chunk = costs_batch[i:i+500]
                self.supabase.table("mp_costs").upsert(
                    chunk, on_conflict="user_id,product_id,marketplace,date,fulfillment_type"
                ).execute()
            records_count = len(costs_batch)

            # Удаляем старые записи деталей WB за период и вставляем новые
            # Удаляем все fulfillment_type (FBO+FBS), т.к. re-sync полностью пересоздаёт данные
            delete_query = (
                self.supabase.table("mp_costs_details")
                .delete()
                .eq("marketplace", "wb")
                .eq("user_id", self.user_id)
                .gte("date", date_from.strftime("%Y-%m-%d"))
                .lte("date", date_to.strftime("%Y-%m-%d"))
            )
            delete_query.execute()

            details_batch = []
            for (pid, date, ft, category, subcategory), data in details_agg.items():
                insert_row = {
                    "product_id": pid,
                    "marketplace": "wb",
                    "date": date,
                    "category": category,
                    "subcategory": subcategory,
                    "amount": round(float(data["amount"] or 0), 2),
                    "operation_type": data.get("operation_type", "WB"),
                    "operation_id": data.get("operation_id", ""),
                    "fulfillment_type": ft,
                }
                if self.user_id:
                    insert_row["user_id"] = self.user_id
                details_batch.append(insert_row)

            for i in range(0, len(details_batch), 500):
                chunk = details_batch[i:i+500]
                self.supabase.table("mp_costs_details").insert(chunk).execute()
            details_count = len(details_batch)

            self._log_sync("wb", "costs", "success", records_count, started_at=started_at)
            return {"status": "success", "records": records_count, "details": details_count}

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Ошибка синхронизации удержаний WB: {error_msg}")
            self._log_sync("wb", "costs", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    @staticmethod
    def _determine_wb_fulfillment(row: dict) -> str:
        """
        Определяет тип фулфилмента из строки WB reportDetailByPeriod.
        Приоритет: isSupply → delivery_type_id → delivery_method → srv_dbs → default FBO.
        WB API: isSupply=true → FBS (продавец отгружает сам).

        Поля по версиям API:
        - isSupply / delivery_type_id: устаревшие, могут отсутствовать в ответе
        - delivery_method: "FBW, (МГТ, коробка)" / "FBS, ..." — присутствует в reportDetailByPeriod
        - srv_dbs: True = Delivery by Seller (≈ FBS) — присутствует в reportDetailByPeriod
        """
        # Priority 1: isSupply (некоторые версии API)
        is_supply = row.get("isSupply")
        if is_supply is True:
            return "FBS"
        if is_supply is False:
            return "FBO"

        # Priority 2: delivery_type_id (некоторые версии API)
        dt_id = row.get("delivery_type_id")
        if dt_id == 2:
            return "FBS"
        if dt_id == 1:
            return "FBO"

        # Priority 3: delivery_method (reportDetailByPeriod — основное поле)
        dm = str(row.get("delivery_method") or "").upper()
        if "FBS" in dm or "DBS" in dm:
            return "FBS"
        if "FBW" in dm or "FBO" in dm:
            return "FBO"

        # Priority 4: srv_dbs — Delivery by Seller (≈ FBS)
        srv_dbs = row.get("srv_dbs")
        if srv_dbs is True:
            return "FBS"

        return "FBO"

    def _classify_wb_report_row(self, row: dict) -> list[dict]:
        """
        Классифицирует строку WB reportDetailByPeriod в записи для mp_costs_details.
        Суммы в details: отрицательная = удержание, положительная = начисление.
        """
        records: list[dict] = []

        def f(key: str) -> float:
            try:
                return float(row.get(key, 0) or 0)
            except Exception:
                return 0.0

        doc = str(row.get("doc_type_name") or "")
        oper = str(row.get("supplier_oper_name") or "")

        # Продажи (WB реализовал Товар (Пр) ≈ retail_amount)
        if doc == "Продажа" and oper == "Продажа":
            sales_amt = f("retail_amount")
            if sales_amt:
                records.append({
                    "category": "Продажи",
                    "subcategory": "Продажа",
                    "amount": sales_amt,
                })

        # Вознаграждение WB (ВВ) и НДС — как в кабинетной выгрузке (две отдельные колонки)
        vw = f("ppvz_vw")
        if vw:
            records.append({
                "category": "Вознаграждение Вайлдберриз (ВВ)",
                "subcategory": "Вознаграждение Вайлдберриз (ВВ), без НДС",
                "amount": vw,  # в отчёте WB уже со знаком
            })
        vw_nds = f("ppvz_vw_nds")
        if vw_nds:
            records.append({
                "category": "Вознаграждение Вайлдберриз (ВВ)",
                "subcategory": "НДС с Вознаграждения Вайлдберриз",
                "amount": vw_nds,  # в отчёте WB уже со знаком
            })

        # Эквайринг
        acq = abs(f("acquiring_fee"))
        if acq:
            sub = str(row.get("payment_processing") or "Эквайринг")
            records.append({
                "category": "Эквайринг/Комиссии за организацию платежей",
                "subcategory": sub,
                "amount": -acq,
            })

        # Услуги доставки товара покупателю (расход)
        delivery = abs(f("delivery_rub"))
        if delivery:
            records.append({
                "category": "Услуги по доставке товара покупателю",
                "subcategory": "Услуги по доставке товара покупателю",
                "amount": -delivery,
            })

        # Возмещения (как в выгрузке) — это начисления, а не удержания.
        pvz = f("ppvz_reward")
        if pvz:
            records.append({
                "category": "Возмещения",
                "subcategory": "Возмещение за выдачу и возврат товаров на ПВЗ",
                "amount": pvz,
            })

        rebill = f("rebill_logistic_cost")
        if rebill:
            records.append({
                "category": "Возмещения",
                "subcategory": "Возмещение издержек по перевозке/по складским операциям с товаром",
                "amount": rebill,
            })

        # Хранение
        storage = abs(f("storage_fee"))
        if storage:
            records.append({
                "category": "Стоимость хранения",
                "subcategory": "Стоимость хранения",
                "amount": -storage,
            })

        # Штрафы
        penalty = abs(f("penalty"))
        if penalty:
            sub = str(row.get("bonus_type_name") or "Штрафы")
            records.append({
                "category": "Общая сумма штрафов",
                "subcategory": sub,
                "amount": -penalty,
            })

        # Прочие удержания/выплаты (кабинетная колонка)
        deduction = abs(f("deduction"))
        if deduction:
            records.append({
                "category": "Прочие удержания/выплаты",
                "subcategory": "Прочие удержания/выплаты",
                "amount": -deduction,
            })

        acceptance = abs(f("acceptance"))
        if acceptance:
            records.append({
                "category": "Стоимость операций при приемке",
                "subcategory": "Стоимость операций при приемке",
                "amount": -acceptance,
            })

        add = f("additional_payment")
        if add:
            records.append({
                "category": "Прочие удержания/выплаты",
                "subcategory": "Прочие удержания/выплаты",
                "amount": add,  # может быть и + и -
            })

        # Лояльность/кэшбэк (поля WB отчёта)
        cashback_amount = f("cashback_amount")
        if cashback_amount:
            records.append({
                "category": "Компенсация скидки по программе лояльности",
                "subcategory": "Компенсация скидки по программе лояльности",
                "amount": cashback_amount,
            })

        cashback_discount = abs(f("cashback_discount"))
        if cashback_discount:
            records.append({
                "category": "Стоимость участия в программе лояльности",
                "subcategory": "Стоимость участия в программе лояльности",
                "amount": -cashback_discount,
            })

        cashback_commission_change = f("cashback_commission_change")
        if cashback_commission_change:
            records.append({
                "category": "Корректировка Вознаграждения Вайлдберриз (ВВ)",
                "subcategory": "Корректировка Вознаграждения Вайлдберриз (ВВ)",
                "amount": cashback_commission_change,
            })

        payment_schedule = f("payment_schedule")
        if payment_schedule:
            records.append({
                "category": "Разовое изменение срока перечисления денежных средств",
                "subcategory": "Разовое изменение срока перечисления денежных средств",
                "amount": payment_schedule,
            })

        return records

    # Маппинг delivery_schema → fulfillment_type (7 значений Ozon API)
    DELIVERY_SCHEMA_MAP = {
        "FBS": "FBS",
        "FBO": "FBO",
        "RFBS": "FBS",        # rFBS — доставка продавцом → аналог FBS
        "FBP": "FBO",         # партнёрские склады → аналог FBO
        "CROSSBORDER": "FBS", # трансграничная → аналог FBS
        "FBOECONOMY": "FBO",  # эконом FBO → FBO
        "FBSECONOMY": "FBS",  # эконом FBS → FBS
    }

    # Ключевые слова для определения логистических сервисов Ozon
    LOGISTICS_SERVICE_KEYWORDS = ("Logistic", "LastMile", "DeliveryCost", "DeliveryKGT")

    # Маппинг operation_type → (category, subcategory) для tree-view
    OZON_COSTS_CATEGORY_MAP = {
        "OperationAgentDeliveredToCustomer": ("Продажи", "Выручка"),
        "OperationAgentDeliveredToCustomerCanceled": ("Продажи", "Корректировка доставки"),
        "OperationItemReturn": ("Услуги доставки", "Возвраты"),
        "MarketplaceRedistributionOfAcquiringOperation": ("Услуги агентов", "Эквайринг"),
        "StarsMembership": ("Услуги агентов", "Звёздные товары"),
        "OperationMarketplaceServicePremiumCashbackIndividualPoints": ("Продвижение и реклама", "Бонусы продавца"),
        "OperationMarketplaceServiceStorage": ("Услуги FBO", "Размещение товаров"),
        "OperationDefectiveWriteOff": ("Компенсации", "Бракованный товар"),
        "OperationLackWriteOff": ("Компенсации", "Недостача товара"),
        "MarketplaceSellerCompensationLossOfGoodsOperation": ("Компенсации", "Потерянный товар"),
        "OperationElectronicServiceStencil": ("Продвижение и реклама", "Трафареты"),
        "OperationElectronicServicesPromotionInSearch": ("Продвижение и реклама", "Продвижение в поиске"),
    }

    def _get_ozon_product_subcategory(self, sku: str) -> str:
        """
        Динамическая классификация товара для costs-tree подкатегории.
        Использует имя товара из mp_products вместо hardcoded SKU списков.
        SKU -> barcode (ozon_sku_map) -> product name (products_map).
        """
        barcode = self.ozon_sku_map.get(str(sku))
        if barcode:
            products_map = self._get_products_map()
            product = products_map.get(barcode, {})
            name = product.get("name", "")
            if name:
                # Обрезаем длинные имена для аккуратного отображения в tree
                return name[:50] if len(name) > 50 else name
        return "Прочее"

    def _classify_ozon_operation(self, op: dict) -> list[dict]:
        """
        Классифицирует операцию Ozon в записи для mp_costs_details.
        Возвращает список записей (одна операция может генерировать несколько записей).
        """
        op_type = op.get("operation_type", "")
        op_type_name = op.get("operation_type_name", "")
        amount = float(op.get("amount", 0))
        sale_commission = float(op.get("sale_commission", 0))
        accruals = float(op.get("accruals_for_sale", 0))
        services = op.get("services", [])
        items = op.get("items", [])

        records = []

        if op_type == "OperationAgentDeliveredToCustomer":
            # Это заказ — разбиваем на: Продажи, Комиссия, Услуги доставки, Услуги агентов.
            # Важно: суммы в finance API часто идут на уровне отправления/заказа,
            # поэтому ниже помечаем записи allocation="order" — дальше они будут
            # распределены между items пропорционально quantity (чтобы не было
            # мультипликации при нескольких товарах в одной операции).
            if accruals:
                records.append({
                    "category": "Продажи",
                    "subcategory": "Выручка",
                    "amount": accruals,
                    "operation_type": op_type,
                    "allocation": "order",
                })
            if sale_commission:
                # Динамическая классификация: используем имя товара из mp_products
                sku = str(items[0].get("sku", "")) if items else ""
                sub = self._get_ozon_product_subcategory(sku)
                records.append({
                    "category": "Вознаграждение Ozon",
                    "subcategory": sub,
                    "amount": sale_commission,  # уже отрицательная
                    "operation_type": op_type,
                    "allocation": "order",
                })
            # Сервисы доставки/агентов (Logistic vs LastMile).
            # В ЛК Ozon "Последняя миля" отображается в "Услуги агентов → Доставка до места выдачи",
            # а не в "Услуги доставки".
            for svc in services:
                svc_name = svc.get("name", "")
                svc_price = float(svc.get("price", 0))
                if svc_price == 0:
                    continue
                if "LastMile" in svc_name:
                    records.append({
                        "category": "Услуги агентов",
                        "subcategory": "Доставка до места выдачи",
                        "amount": svc_price,
                        "operation_type": op_type,
                        "allocation": "order",
                    })
                elif any(kw in svc_name for kw in self.LOGISTICS_SERVICE_KEYWORDS):
                    records.append({
                        "category": "Услуги доставки",
                        "subcategory": "Логистика",
                        "amount": svc_price,
                        "operation_type": op_type,
                        "allocation": "order",
                    })

        elif op_type == "OperationItemReturn":
            # Возврат — сервисы логистики возврата
            for svc in services:
                svc_price = float(svc.get("price", 0))
                if svc_price == 0:
                    continue
                records.append({
                    "category": "Услуги доставки",
                    "subcategory": "Возвраты",
                    "amount": svc_price,
                    "operation_type": op_type,
                    "allocation": "order",
                })

        elif op_type == "MarketplaceRedistributionOfAcquiringOperation":
            records.append({
                "category": "Услуги агентов",
                "subcategory": "Эквайринг",
                "amount": amount,
                "operation_type": op_type,
                "allocation": "order",
            })

        elif op_type == "StarsMembership":
            records.append({
                "category": "Услуги агентов",
                "subcategory": "Звёздные товары",
                "amount": amount,
                "operation_type": op_type,
                "allocation": "order",
            })

        elif op_type == "OperationMarketplaceServicePremiumCashbackIndividualPoints":
            records.append({
                "category": "Продвижение и реклама",
                "subcategory": "Бонусы продавца",
                "amount": amount,
                "operation_type": op_type,
                "allocation": "order",
            })

        elif op_type == "OperationMarketplaceServiceStorage":
            records.append({
                "category": "Услуги FBO",
                "subcategory": "Размещение товаров",
                "amount": amount,
                "operation_type": op_type,
                "allocation": "order",
            })

        # --- Новые operation_types (Phase 6) ---

        elif op_type == "OperationAgentDeliveredToCustomerCanceled":
            # Корректировка/отмена доставки — реверс начисления
            if amount != 0:
                records.append({
                    "category": "Продажи",
                    "subcategory": "Корректировка доставки",
                    "amount": amount,
                    "operation_type": op_type,
                    "allocation": "order",
                })
            # Сервисы логистики (реверс)
            for svc in services:
                svc_price = float(svc.get("price", 0))
                if svc_price == 0:
                    continue
                svc_name = svc.get("name", "")
                if "LastMile" in svc_name:
                    records.append({
                        "category": "Услуги агентов",
                        "subcategory": "Доставка до места выдачи",
                        "amount": svc_price,
                        "operation_type": op_type,
                        "allocation": "order",
                    })
                elif any(kw in svc_name for kw in self.LOGISTICS_SERVICE_KEYWORDS):
                    records.append({
                        "category": "Услуги доставки",
                        "subcategory": "Логистика",
                        "amount": svc_price,
                        "operation_type": op_type,
                        "allocation": "order",
                    })
                else:
                    records.append({
                        "category": "Прочее",
                        "subcategory": svc_name or "Прочие услуги",
                        "amount": svc_price,
                        "operation_type": op_type,
                        "allocation": "order",
                    })

        elif op_type in ("OperationDefectiveWriteOff", "OperationLackWriteOff",
                         "MarketplaceSellerCompensationLossOfGoodsOperation"):
            # Компенсации за бракованный/недостающий/потерянный товар
            subcategory_map = {
                "OperationDefectiveWriteOff": "Бракованный товар",
                "OperationLackWriteOff": "Недостача товара",
                "MarketplaceSellerCompensationLossOfGoodsOperation": "Потерянный товар",
            }
            if amount != 0:
                records.append({
                    "category": "Компенсации",
                    "subcategory": subcategory_map[op_type],
                    "amount": amount,
                    "operation_type": op_type,
                    "allocation": "order",
                })

        elif op_type in ("OperationElectronicServiceStencil",
                         "OperationElectronicServicesPromotionInSearch"):
            # Рекламные услуги (трафареты, продвижение в поиске)
            subcategory_map = {
                "OperationElectronicServiceStencil": "Трафареты",
                "OperationElectronicServicesPromotionInSearch": "Продвижение в поиске",
            }
            if amount != 0:
                records.append({
                    "category": "Продвижение и реклама",
                    "subcategory": subcategory_map[op_type],
                    "amount": amount,
                    "operation_type": op_type,
                    "allocation": "order",
                })

        else:
            # Неизвестный тип — сохраняем как "Прочее"
            if amount != 0:
                records.append({
                    "category": "Прочее",
                    "subcategory": op_type_name or op_type,
                    "amount": amount,
                    "operation_type": op_type,
                })

        return records

    async def sync_costs_ozon(self, date_from: datetime, date_to: datetime = None) -> dict:
        """Синхронизация удержаний с Ozon (mp_costs + mp_costs_details)"""
        started_at = datetime.now()
        date_to = date_to or datetime.now()
        records_count = 0
        details_count = 0

        try:
            products_map = self._get_products_map()

            # Ozon finance API requires period within a single calendar month.
            # Split into calendar-month chunks (e.g., Feb 22→Feb 28, Mar 1→Mar 24).
            all_operations = []
            chunk_from = date_from
            chunk_count = 0
            max_chunks = 14  # safety limit: ~1 year

            while chunk_from < date_to and chunk_count < max_chunks:
                # End of current month
                if chunk_from.month == 12:
                    month_end = chunk_from.replace(year=chunk_from.year + 1, month=1, day=1) - timedelta(days=1)
                else:
                    month_end = chunk_from.replace(month=chunk_from.month + 1, day=1) - timedelta(days=1)
                chunk_to = min(month_end, date_to)

                # Получаем транзакции для текущего чанка с пагинацией
                page = 1
                while True:
                    result = await self.ozon_client.get_finance_transaction_list(chunk_from, chunk_to, page=page)
                    operations = result.get("result", {}).get("operations", [])
                    if not operations:
                        break
                    all_operations.extend(operations)
                    page += 1
                    if page > 20:  # safety limit per chunk
                        break

                logger.info(f"Ozon finance chunk {chunk_count + 1}: {len(all_operations)} ops total, period {chunk_from.strftime('%Y-%m-%d')} - {chunk_to.strftime('%Y-%m-%d')}")
                chunk_from = chunk_to + timedelta(days=1)
                chunk_count += 1

            logger.info(f"Ozon finance: получено {len(all_operations)} операций за {date_from.strftime('%Y-%m-%d')} - {date_to.strftime('%Y-%m-%d')} ({chunk_count} chunks)")

            # === 1. Агрегация для mp_costs ===
            costs_agg = {}  # {(barcode, date, ft): {commission, logistics, ...}}

            for op in all_operations:
                date = op.get("operation_date", "")[:10]
                items = op.get("items", [])
                op_type = op.get("operation_type", "")
                amount = float(op.get("amount", 0))
                sale_commission = float(op.get("sale_commission", 0))
                services = op.get("services", [])

                # Определяем FBO/FBS из posting.delivery_schema (7 значений)
                posting = op.get("posting", {}) or {}
                delivery_schema = str(posting.get("delivery_schema") or "").upper()
                ft = self.DELIVERY_SCHEMA_MAP.get(delivery_schema, "FBO")

                # FIX: OperationMarketplaceServiceStorage может приходить без items
                # (storage charges — account-level, без привязки к товару).
                # Распределяем по всем товарам с Ozon costs за этот день пропорционально,
                # или если items есть — маппим напрямую.
                if op_type == "OperationMarketplaceServiceStorage" and not items:
                    # Storage без items: распределяем сумму по ВСЕМ ключам за этот день
                    # Если ключей ещё нет — создаём "placeholder" для первого Ozon товара
                    storage_amount = abs(amount)
                    day_keys = [k for k in costs_agg if k[1] == date and k[2] == ft]
                    if day_keys:
                        # Распределяем поровну между товарами (storage не per-item)
                        share = storage_amount / len(day_keys)
                        for dk in day_keys:
                            costs_agg[dk]["storage"] += share
                    else:
                        # Нет ключей за этот день — сохраняем в первый Ozon товар как fallback
                        first_barcode = next(
                            (bc for bc in products_map if products_map[bc].get("ozon_product_id")),
                            None,
                        )
                        if first_barcode:
                            key = (first_barcode, date, ft)
                            if key not in costs_agg:
                                costs_agg[key] = {
                                    "commission": 0, "logistics": 0, "storage": 0,
                                    "promotion": 0, "penalties": 0, "acquiring": 0, "other": 0,
                                    "settled_qty": 0,
                                }
                            costs_agg[key]["storage"] += storage_amount
                    continue

                for item in items:
                    sku = str(item.get("sku", ""))
                    barcode = self.ozon_sku_map.get(sku)
                    if not barcode:
                        continue

                    key = (barcode, date, ft)
                    if key not in costs_agg:
                        costs_agg[key] = {
                            "commission": 0, "logistics": 0, "storage": 0,
                            "promotion": 0, "penalties": 0, "acquiring": 0, "other": 0,
                            "settled_qty": 0,
                        }

                    if op_type == "OperationAgentDeliveredToCustomer":
                        # Считаем settled_qty — количество единиц, проданных по дате расчёта
                        try:
                            item_qty = int(item.get("quantity", 1) or 1)
                        except Exception:
                            item_qty = 1
                        if item_qty <= 0:
                            item_qty = 1
                        costs_agg[key]["settled_qty"] += item_qty

                        costs_agg[key]["commission"] += abs(sale_commission)
                        for svc in services:
                            svc_name = svc.get("name", "")
                            svc_price = abs(float(svc.get("price", 0)))
                            if any(kw in svc_name for kw in self.LOGISTICS_SERVICE_KEYWORDS):
                                costs_agg[key]["logistics"] += svc_price

                    elif op_type == "OperationAgentDeliveredToCustomerCanceled":
                        # Реверс доставки — ВЫЧИТАЕМ ранее начисленные удержания
                        costs_agg[key]["commission"] -= abs(sale_commission)
                        for svc in services:
                            svc_name = svc.get("name", "")
                            svc_price = abs(float(svc.get("price", 0)))
                            if any(kw in svc_name for kw in self.LOGISTICS_SERVICE_KEYWORDS):
                                costs_agg[key]["logistics"] -= svc_price
                            else:
                                costs_agg[key]["other"] -= svc_price

                    elif op_type == "OperationItemReturn":
                        for svc in services:
                            costs_agg[key]["logistics"] += abs(float(svc.get("price", 0)))
                    elif op_type == "MarketplaceRedistributionOfAcquiringOperation":
                        costs_agg[key]["acquiring"] += abs(amount)
                    elif op_type == "OperationMarketplaceServiceStorage":
                        costs_agg[key]["storage"] += abs(amount)
                    elif op_type == "OperationMarketplaceServicePremiumCashbackIndividualPoints":
                        costs_agg[key]["promotion"] += abs(amount)
                    elif op_type in ("OperationElectronicServiceStencil",
                                     "OperationElectronicServicesPromotionInSearch"):
                        # Рекламные услуги — трафареты, продвижение в поиске
                        costs_agg[key]["promotion"] += abs(amount)
                    elif op_type in ("OperationDefectiveWriteOff", "OperationLackWriteOff",
                                     "MarketplaceSellerCompensationLossOfGoodsOperation"):
                        # Компенсации за товар
                        costs_agg[key]["other"] += abs(amount)
                    elif op_type == "StarsMembership":
                        costs_agg[key]["other"] += abs(amount)
                    else:
                        costs_agg[key]["other"] += abs(amount)

            # Сохраняем mp_costs (batch upsert по 500)
            costs_batch = []
            for (barcode, date, ft), costs in costs_agg.items():
                product_id = products_map.get(barcode, {}).get("id")
                if not product_id:
                    continue

                upsert_row = {
                    "product_id": product_id,
                    "marketplace": "ozon",
                    "date": date,
                    "commission": costs["commission"],
                    "logistics": costs["logistics"],
                    "storage": costs["storage"],
                    "promotion": costs["promotion"],
                    "penalties": costs["penalties"],
                    "acquiring": costs["acquiring"],
                    "other_costs": costs["other"],
                    "fulfillment_type": ft,
                    "settled_qty": costs.get("settled_qty", 0),
                }
                if self.user_id:
                    upsert_row["user_id"] = self.user_id
                costs_batch.append(upsert_row)

            for i in range(0, len(costs_batch), 500):
                chunk = costs_batch[i:i+500]
                self.supabase.table("mp_costs").upsert(
                    chunk, on_conflict="user_id,product_id,marketplace,date,fulfillment_type"
                ).execute()
            records_count = len(costs_batch)

            # === 2. Гранулярные данные для mp_costs_details (tree-view) ===
            details_agg = {}  # {(product_id, date, ft, category, subcategory, order_date): {amount, ...}}

            for op in all_operations:
                date = op.get("operation_date", "")[:10]
                items = op.get("items", [])
                operation_id = str(op.get("operation_id", ""))

                # Определяем FBO/FBS (7 значений delivery_schema)
                posting = op.get("posting", {}) or {}
                delivery_schema = str(posting.get("delivery_schema") or "").upper()
                ft = self.DELIVERY_SCHEMA_MAP.get(delivery_schema, "FBO")

                # order_date: дата принятия заказа в обработку (из posting.order_date)
                # Используется в UE endpoint для фильтрации по дате заказа (не settlement)
                order_date = str(posting.get("order_date") or "")[:10] or None

                # posting_number for exact JOIN with mp_orders (Rule #49: DATA-002 fix)
                raw_posting_number = str(posting.get("posting_number") or "").strip() or None

                detail_records = self._classify_ozon_operation(op)

                if items:
                    # Распределяем "order-level" суммы по товарам, чтобы не было мультипликации
                    # при нескольких items в одной операции.
                    total_qty = 0
                    item_qtys = []
                    for it in items:
                        try:
                            q = int(it.get("quantity", 1) or 1)
                        except Exception:
                            q = 1
                        if q <= 0:
                            q = 1
                        item_qtys.append(q)
                        total_qty += q
                    if total_qty <= 0:
                        total_qty = len(items)

                    for item in items:
                        sku = str(item.get("sku", ""))
                        barcode = self.ozon_sku_map.get(sku)
                        if not barcode:
                            continue
                        pid = products_map.get(barcode, {}).get("id")
                        if not pid:
                            continue

                        # posting_number matching mp_orders.order_id format (Rule #49)
                        item_posting_number = None
                        if raw_posting_number:
                            if len(items) > 1 and sku:
                                item_posting_number = f"{raw_posting_number}_{sku}"
                            else:
                                item_posting_number = raw_posting_number

                        # share by quantity (fallback: equal split)
                        try:
                            q = int(item.get("quantity", 1) or 1)
                        except Exception:
                            q = 1
                        if q <= 0:
                            q = 1
                        share = (q / total_qty) if total_qty > 0 else (1.0 / len(items))

                        for rec in detail_records:
                            alloc = rec.get("allocation")
                            amt = rec["amount"] * share if alloc == "order" else rec["amount"]
                            key = (pid, date, ft, rec["category"], rec["subcategory"], order_date, item_posting_number)
                            if key not in details_agg:
                                details_agg[key] = {
                                    "amount": 0,
                                    "operation_type": rec["operation_type"],
                                    "operation_id": operation_id,
                                }
                            details_agg[key]["amount"] += amt
                else:
                    # Операции без привязки к товару (хранение и т.п.)
                    # Распределяем равномерно между всеми товарами с продажами на Ozon
                    ozon_product_ids = [
                        p["id"] for p in products_map.values()
                        if p.get("ozon_product_id")
                    ]
                    if not ozon_product_ids:
                        continue
                    # Важно: делим в копейках так, чтобы сумма распределённых значений
                    # совпадала с исходной суммой до 0.01 ₽ (без накопления ошибок округления).
                    ozon_product_ids = sorted(ozon_product_ids)

                    for rec in detail_records:
                        total_cents = int(round(float(rec["amount"]) * 100))
                        n = len(ozon_product_ids)
                        if n == 0:
                            continue
                        sign = 1 if total_cents >= 0 else -1
                        abs_cents = abs(total_cents)
                        base_abs = abs_cents // n
                        rem = abs_cents % n

                        for i, pid in enumerate(ozon_product_ids):
                            key = (pid, date, ft, rec["category"], rec["subcategory"], order_date, None)
                            if key not in details_agg:
                                details_agg[key] = {
                                    "amount": 0,
                                    "operation_type": rec["operation_type"],
                                    "operation_id": operation_id,
                                }
                            cents = base_abs + (1 if i < rem else 0)
                            details_agg[key]["amount"] += sign * (cents / 100.0)

            # Удаляем старые записи за период перед вставкой
            # Удаляем все fulfillment_type (FBO+FBS), т.к. re-sync полностью пересоздаёт данные
            delete_query = (
                self.supabase.table("mp_costs_details")
                .delete()
                .eq("marketplace", "ozon")
                .eq("user_id", self.user_id)
                .gte("date", date_from.strftime("%Y-%m-%d"))
                .lte("date", date_to.strftime("%Y-%m-%d"))
            )
            delete_query.execute()

            # Сохраняем mp_costs_details (batch insert по 500)
            details_batch = []
            for (pid, date, ft, category, subcategory, od, pn), data in details_agg.items():
                insert_row = {
                    "product_id": pid,
                    "marketplace": "ozon",
                    "date": date,
                    "category": category,
                    "subcategory": subcategory,
                    "amount": round(data["amount"], 2),
                    "operation_type": data["operation_type"],
                    "operation_id": data["operation_id"],
                    "fulfillment_type": ft,
                }
                if od:
                    insert_row["order_date"] = od
                if pn:
                    insert_row["posting_number"] = pn
                if self.user_id:
                    insert_row["user_id"] = self.user_id
                details_batch.append(insert_row)

            for i in range(0, len(details_batch), 500):
                chunk = details_batch[i:i+500]
                self.supabase.table("mp_costs_details").insert(chunk).execute()
            details_count = len(details_batch)

            logger.info(f"Ozon costs: {records_count} mp_costs records, {details_count} mp_costs_details records")
            self._log_sync("ozon", "costs", "success", records_count, started_at=started_at)
            return {"status": "success", "records": records_count, "details": details_count}

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Ошибка синхронизации удержаний Ozon: {error_msg}")
            self._log_sync("ozon", "costs", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    # ==================== СИНХРОНИЗАЦИЯ РЕКЛАМЫ ====================

    async def sync_ads_wb(self, date_from: datetime, date_to: datetime = None) -> dict:
        """Синхронизация рекламных расходов с Wildberries"""
        started_at = datetime.now()
        date_to = date_to or datetime.now()
        records_count = 0

        try:
            products_map = self._get_products_map()

            # Получаем кампании
            campaigns = await self.wb_client.get_advert_campaigns()
            campaign_ids = []
            for adv_type in campaigns.get("adverts", []):
                for ad in adv_type.get("advert_list", []):
                    if ad.get("advertId"):
                        campaign_ids.append(ad["advertId"])

            if not campaign_ids:
                self._log_sync("wb", "ads", "success", 0, started_at=started_at)
                return {"status": "success", "records": 0}

            # Получаем статистику по одной кампании (WB API ограничение + rate limit)
            import asyncio
            stats = []
            for cid in campaign_ids:
                try:
                    result = await self.wb_client.get_advert_stats([cid], date_from, date_to)
                    if result and isinstance(result, list):
                        stats.extend(result)
                except Exception as e:
                    logger.warning(f"WB Ads: не удалось получить статистику кампании {cid}: {e}")
                await asyncio.sleep(5)  # Rate limit: WB Ads API ограничивает частоту

            ads_batch = []
            for campaign_stat in stats:
                campaign_id = campaign_stat.get("advertId")

                for day in campaign_stat.get("days", []):
                    date = day.get("date")

                    # Aggregate across all appTypes per nmId
                    # WB returns separate entries per appType (search/catalog/card)
                    # but our unique key is (user,product,mp,date,campaign) — need totals
                    nm_totals: dict[int, dict] = {}
                    for app in day.get("apps", []):
                        for nm in app.get("nm", []):
                            nm_id = nm.get("nmId")
                            if not nm_id:
                                continue
                            if nm_id not in nm_totals:
                                nm_totals[nm_id] = {
                                    "views": 0, "clicks": 0,
                                    "cost": 0.0, "orders": 0,
                                    "orders_sum_rub": 0.0,
                                }
                            t = nm_totals[nm_id]
                            t["views"] += nm.get("views", 0)
                            t["clicks"] += nm.get("clicks", 0)
                            t["cost"] += float(nm.get("sum", 0))
                            t["orders"] += nm.get("orders", 0)
                            t["orders_sum_rub"] += float(nm.get("ordersSumRub", 0) or 0)

                    for nm_id, t in nm_totals.items():
                        # Находим товар
                        product_id = None
                        for barcode, product in products_map.items():
                            if product.get("wb_nm_id") == nm_id:
                                product_id = product["id"]
                                break

                        if not product_id:
                            continue

                        views = t["views"]
                        clicks = t["clicks"]
                        cost = t["cost"]
                        orders = t["orders"]

                        ctr = round(clicks / views * 100, 2) if views > 0 else 0
                        cpc = round(cost / clicks, 2) if clicks > 0 else 0
                        acos = round(cost / t["orders_sum_rub"] * 100, 2) if t["orders_sum_rub"] > 0 else None

                        upsert_row = {
                            "product_id": product_id,
                            "marketplace": "wb",
                            "date": date,
                            "campaign_id": str(campaign_id),
                            "campaign_name": str(campaign_id),
                            "impressions": views,
                            "clicks": clicks,
                            "cost": cost,
                            "orders_count": orders,
                            "ctr": ctr,
                            "cpc": cpc,
                            "acos": acos,
                        }
                        if self.user_id:
                            upsert_row["user_id"] = self.user_id
                        ads_batch.append(upsert_row)

            # Batch upsert по 500 (паттерн из sync_costs_wb)
            for i in range(0, len(ads_batch), 500):
                chunk = ads_batch[i:i+500]
                self.supabase.table("mp_ad_costs").upsert(
                    chunk, on_conflict="user_id,product_id,marketplace,date,campaign_id"
                ).execute()
            records_count = len(ads_batch)

            self._log_sync("wb", "ads", "success", records_count, started_at=started_at)
            return {"status": "success", "records": records_count}

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Ошибка синхронизации рекламы WB: {error_msg}")
            self._log_sync("wb", "ads", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    async def sync_ads_ozon(self, date_from: datetime, date_to: datetime = None) -> dict:
        """Синхронизация рекламных расходов с Ozon"""
        import asyncio
        started_at = datetime.now()
        date_to = date_to or datetime.now()
        records_count = 0

        try:
            products_map = self._get_products_map()

            # Получаем кампании (только SKU-типа, не REF_VK/SEARCH_PROMO)
            campaigns = await self.ozon_perf_client.get_campaigns()
            sku_campaigns = [
                c for c in campaigns.get("list", [])
                if c.get("id") and c.get("advObjectType") == "SKU"
            ]

            if not sku_campaigns:
                self._log_sync("ozon", "ads", "success", 0, started_at=started_at)
                return {"status": "success", "records": 0}

            # DELETE перед INSERT: Ozon ads имеют product_id=None (account-level).
            # PostgreSQL: NULL != NULL в UNIQUE constraints → UPSERT всегда INSERT → дупликация.
            # Решение: удаляем старые записи за период, затем вставляем новые.
            delete_query = (
                self.supabase.table("mp_ad_costs")
                .delete()
                .eq("marketplace", "ozon")
                .eq("user_id", self.user_id)
                .gte("date", date_from.strftime("%Y-%m-%d"))
                .lte("date", date_to.strftime("%Y-%m-%d"))
            )
            delete_query.execute()

            # Обрабатываем по одной кампании (ограничение API: 1 запрос одновременно)
            ads_batch = []
            for campaign in sku_campaigns:
                campaign_id = str(campaign["id"])
                try:
                    stats = await self.ozon_perf_client.get_campaign_stats(
                        [campaign_id], date_from, date_to
                    )

                    for row in stats.get("rows", []):
                        date = row.get("date")
                        if not date:
                            continue

                        views = row.get("views", 0)
                        clicks = row.get("clicks", 0)
                        cost = float(row.get("expense", 0))
                        orders = row.get("orders", 0)
                        ad_revenue = float(row.get("ad_revenue", 0))

                        # Пропускаем дни без активности
                        if views == 0 and clicks == 0 and cost == 0:
                            continue

                        ctr = round(clicks / views * 100, 2) if views > 0 else 0
                        cpc = round(cost / clicks, 2) if clicks > 0 else 0

                        insert_row = {
                            "product_id": None,
                            "marketplace": "ozon",
                            "date": date,
                            "campaign_id": campaign_id,
                            "campaign_name": campaign.get("title", str(campaign_id)),
                            "impressions": views,
                            "clicks": clicks,
                            "cost": cost,
                            "orders_count": orders,
                            "ctr": ctr,
                            "cpc": cpc,
                            "ad_revenue": ad_revenue,
                        }
                        if self.user_id:
                            insert_row["user_id"] = self.user_id
                        ads_batch.append(insert_row)

                except Exception as e:
                    logger.warning(f"Ozon Ads: не удалось получить статистику кампании {campaign_id}: {e}")

                # Ждём между кампаниями (ограничение: 1 активный отчёт)
                await asyncio.sleep(5)

            # Batch INSERT по 500 (паттерн из sync_ads_wb)
            for i in range(0, len(ads_batch), 500):
                chunk = ads_batch[i:i+500]
                self.supabase.table("mp_ad_costs").insert(chunk).execute()
            records_count = len(ads_batch)

            self._log_sync("ozon", "ads", "success", records_count, started_at=started_at)
            return {"status": "success", "records": records_count}

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Ошибка синхронизации рекламы Ozon: {error_msg}")
            self._log_sync("ozon", "ads", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    # ==================== СИНХРОНИЗАЦИЯ ЗАКАЗОВ (Phase 2 Order Monitor) ====================

    async def sync_orders_wb(self, date_from: datetime, date_to: datetime = None) -> dict:
        """
        Синхронизация позаказных данных WB.
        3-шаговое обогащение через srid:
        1. get_orders() → базовые данные (srid, barcode, price, region, warehouse)
        2. get_sales() → статусы (saleID: S=sold, R=returned, forPay)
        3. get_report_detail() → финансы (commission, logistics, storage, payout)
        """
        import asyncio
        started_at = datetime.now()
        date_to = date_to or datetime.now()
        records_count = 0

        try:
            products_map = self._get_products_map()

            # Шаг 1: Получаем заказы (srid → базовые данные)
            orders_data = await self.wb_client.get_orders(date_from, flag=0)
            if orders_data is None:
                orders_data = []

            # Индексируем заказы по srid
            orders_by_srid: dict[str, dict] = {}
            for order in orders_data:
                srid = order.get("srid")
                if not srid:
                    continue
                barcode = str(order.get("barcode", "")) or ""
                nm_id = order.get("nmId")

                # Мэппинг barcode → product_id
                product_id = None
                if barcode and barcode in products_map:
                    product_id = products_map[barcode]["id"]
                elif nm_id:
                    for bc, p in products_map.items():
                        if p.get("wb_nm_id") == nm_id:
                            product_id = p["id"]
                            barcode = bc
                            break

                orders_by_srid[srid] = {
                    "marketplace": "wb",
                    "order_id": srid,
                    "product_id": product_id,
                    "barcode": barcode,
                    "order_date": order.get("date") or order.get("createdAt") or started_at.isoformat(),
                    "last_change_date": order.get("lastChangeDate"),
                    "status": "cancelled" if order.get("isCancel") else "ordered",
                    "price": float(order.get("totalPrice", 0) or 0),
                    "region": order.get("regionName"),
                    "warehouse": order.get("warehouseName"),
                    "settled": False,
                }

            await asyncio.sleep(1)  # Rate limit между вызовами WB API

            # Шаг 2: Обогащаем статусами из sales (S=sold, R=returned)
            sales_data = await self.wb_client.get_sales(date_from, flag=0)
            if sales_data is None:
                sales_data = []

            for sale in sales_data:
                srid = sale.get("srid")
                if not srid:
                    continue

                sale_id = sale.get("saleID", "")
                for_pay = float(sale.get("forPay", 0) or 0)

                if srid in orders_by_srid:
                    if sale_id.startswith("S"):
                        orders_by_srid[srid]["status"] = "sold"
                    elif sale_id.startswith("R"):
                        orders_by_srid[srid]["status"] = "returned"
                    orders_by_srid[srid]["sale_amount"] = for_pay
                    orders_by_srid[srid]["wb_sale_id"] = sale_id
                    if sale.get("lastChangeDate"):
                        orders_by_srid[srid]["last_change_date"] = sale["lastChangeDate"]
                    # priceWithDisc — цена после скидки СПП (реальная цена продажи)
                    price_with_disc = float(sale.get("priceWithDisc", 0) or 0)
                    if price_with_disc:
                        orders_by_srid[srid]["sale_price"] = price_with_disc
                    # Подтянуть каталожную цену из sales если в orders была 0
                    if price_with_disc and not orders_by_srid[srid].get("price"):
                        orders_by_srid[srid]["price"] = price_with_disc
                else:
                    # Продажа без парного заказа (заказ старше 90 дней) — создаём запись
                    barcode = str(sale.get("barcode", "")) or ""
                    nm_id = sale.get("nmId")
                    product_id = None
                    if barcode and barcode in products_map:
                        product_id = products_map[barcode]["id"]
                    elif nm_id:
                        for bc, p in products_map.items():
                            if p.get("wb_nm_id") == nm_id:
                                product_id = p["id"]
                                barcode = bc
                                break

                    status = "sold" if sale_id.startswith("S") else "returned" if sale_id.startswith("R") else "ordered"
                    price_with_disc = float(sale.get("priceWithDisc", 0) or 0)
                    orders_by_srid[srid] = {
                        "marketplace": "wb",
                        "order_id": srid,
                        "product_id": product_id,
                        "barcode": barcode,
                        "order_date": sale.get("date") or started_at.isoformat(),
                        "last_change_date": sale.get("lastChangeDate"),
                        "status": status,
                        "price": price_with_disc,
                        "sale_price": price_with_disc,
                        "sale_amount": for_pay,
                        "wb_sale_id": sale_id,
                        "region": sale.get("regionName"),
                        "warehouse": sale.get("warehouseName"),
                        "settled": False,
                    }

            await asyncio.sleep(1)  # Rate limit

            # Шаг 3: Обогащаем финансами из reportDetail
            # ВАЖНО: reportDetail возвращает НЕСКОЛЬКО строк на один srid
            # (продажа, логистика, хранение и т.д.) — нужно НАКОПИТЬ значения
            report = await self.wb_client.get_report_detail(date_from, date_to)
            if report is None:
                report = []

            for row in report:
                srid = row.get("srid")
                if not srid:
                    continue

                commission = abs(float(row.get("ppvz_vw", 0) or 0)) + abs(float(row.get("ppvz_vw_nds", 0) or 0))
                logistics_cost = abs(float(row.get("delivery_rub", 0) or 0))
                storage = abs(float(row.get("storage_fee", 0) or 0))
                other = abs(float(row.get("penalty", 0) or 0)) + abs(float(row.get("deduction", 0) or 0)) + abs(float(row.get("acceptance", 0) or 0))
                payout_val = float(row.get("ppvz_for_pay", 0) or 0)
                retail_price = float(row.get("retail_price", 0) or 0)
                retail_price_withdisc = float(row.get("retail_price_withdisc_rub", 0) or 0)
                rrd_id = row.get("rrd_id")
                sale_dt = row.get("sale_dt")  # Дата выкупа (когда покупатель забрал товар)
                doc_type = row.get("doc_type_name", "")

                ft = self._determine_wb_fulfillment(row)

                if srid in orders_by_srid:
                    # Накапливаем финансовые данные (несколько строк на srid)
                    orders_by_srid[srid]["commission"] = orders_by_srid[srid].get("commission", 0) + commission
                    orders_by_srid[srid]["logistics"] = orders_by_srid[srid].get("logistics", 0) + logistics_cost
                    orders_by_srid[srid]["storage_fee"] = orders_by_srid[srid].get("storage_fee", 0) + storage
                    orders_by_srid[srid]["other_fees"] = orders_by_srid[srid].get("other_fees", 0) + other
                    orders_by_srid[srid]["payout"] = (orders_by_srid[srid].get("payout") or 0) + payout_val
                    orders_by_srid[srid]["wb_rrd_id"] = rrd_id
                    orders_by_srid[srid]["settled"] = True
                    orders_by_srid[srid]["_fulfillment"] = ft
                    # sale_dt = дата выкупа. Берём первый ненулевой для "Продажа"
                    if sale_dt and "Возврат" not in doc_type and not orders_by_srid[srid].get("_sale_dt"):
                        orders_by_srid[srid]["_sale_dt"] = sale_dt
                    # retail_price_withdisc_rub — реальная цена после СПП (не накапливаем, берём первое ненулевое)
                    if retail_price_withdisc and not orders_by_srid[srid].get("sale_price"):
                        orders_by_srid[srid]["sale_price"] = retail_price_withdisc
                    # Каталожная цена (до СПП) — fallback
                    if not orders_by_srid[srid].get("price") and retail_price:
                        orders_by_srid[srid]["price"] = retail_price
                else:
                    # Финансовый отчёт без парного заказа/продажи — создаём запись
                    nm_id = row.get("nm_id")
                    barcode = ""
                    product_id = None
                    if nm_id:
                        for bc, p in products_map.items():
                            if p.get("wb_nm_id") == nm_id:
                                product_id = p["id"]
                                barcode = bc
                                break

                    status = "sold"
                    if "Возврат" in doc_type:
                        status = "returned"

                    orders_by_srid[srid] = {
                        "marketplace": "wb",
                        "order_id": srid,
                        "product_id": product_id,
                        "barcode": barcode,
                        "order_date": row.get("rr_dt") or started_at.isoformat(),
                        "status": status,
                        "price": retail_price,
                        "sale_price": retail_price_withdisc if retail_price_withdisc else None,
                        "commission": commission,
                        "logistics": logistics_cost,
                        "storage_fee": storage,
                        "other_fees": other,
                        "payout": payout_val,
                        "wb_rrd_id": rrd_id,
                        "settled": True,
                        "_fulfillment": ft,
                        "_sale_dt": sale_dt if (sale_dt and "Возврат" not in doc_type) else None,
                    }

            # Batch upsert в mp_orders
            batch = []
            for srid, order_data in orders_by_srid.items():
                row = {
                    "marketplace": order_data["marketplace"],
                    "order_id": order_data["order_id"],
                    "product_id": order_data.get("product_id"),
                    "barcode": order_data.get("barcode"),
                    "order_date": order_data["order_date"],
                    "last_change_date": order_data.get("last_change_date"),
                    "status": order_data["status"],
                    "price": order_data.get("price", 0),
                    "sale_price": order_data.get("sale_price"),
                    "sale_amount": order_data.get("sale_amount"),
                    "commission": order_data.get("commission", 0),
                    "logistics": order_data.get("logistics", 0),
                    "storage_fee": order_data.get("storage_fee", 0),
                    "other_fees": order_data.get("other_fees", 0),
                    "payout": order_data.get("payout"),
                    "settled": order_data.get("settled", False),
                    "wb_sale_id": order_data.get("wb_sale_id"),
                    "wb_rrd_id": order_data.get("wb_rrd_id"),
                    "region": order_data.get("region"),
                    "warehouse": order_data.get("warehouse"),
                    "updated_at": datetime.now().isoformat(),
                    "fulfillment_type": order_data.get("_fulfillment", "FBO"),
                    "delivery_date": order_data.get("_sale_dt"),
                }
                if self.user_id:
                    row["user_id"] = self.user_id
                batch.append(row)

            # Upsert пачками по 500
            for i in range(0, len(batch), 500):
                chunk = batch[i:i+500]
                self.supabase.table("mp_orders").upsert(
                    chunk, on_conflict="user_id,marketplace,order_id"
                ).execute()
                records_count += len(chunk)

            self._log_sync("wb", "orders", "success", records_count, started_at=started_at)
            return {"status": "success", "records": records_count}

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Ошибка синхронизации заказов WB: {error_msg}")
            self._log_sync("wb", "orders", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    async def sync_orders_ozon(self, date_from: datetime, date_to: datetime = None) -> dict:
        """
        Синхронизация позаказных данных Ozon.
        Запрашивает FBS и FBO отправления с financial_data.
        """
        started_at = datetime.now()
        date_to = date_to or datetime.now()
        records_count = 0

        try:
            products_map = self._get_products_map()
            ozon_sku_map = self.ozon_sku_map

            all_postings = []

            # FBS отправления
            offset = 0
            limit = 1000
            while True:
                result = await self.ozon_client.get_posting_fbs_list(date_from, date_to, limit=limit, offset=offset)
                postings = result.get("result", {}).get("postings", [])
                if not postings:
                    break
                for p in postings:
                    p["_fulfillment"] = "FBS"
                all_postings.extend(postings)
                if len(postings) < limit:
                    break
                offset += limit

            # FBO отправления
            offset = 0
            while True:
                result = await self.ozon_client.get_posting_fbo_list(date_from, date_to, limit=limit, offset=offset)
                postings = result.get("result", [])
                if not postings:
                    break
                for p in postings:
                    p["_fulfillment"] = "FBO"
                all_postings.extend(postings)
                if len(postings) < limit:
                    break
                offset += limit

            logger.info(f"Ozon orders: {len(all_postings)} postings (FBS+FBO)")

            # Маппинг статусов Ozon → наши статусы
            STATUS_MAP = {
                # Общие для FBS и FBO (5 статусов FBO)
                "delivered": "sold",
                "cancelled": "cancelled",
                "awaiting_packaging": "ordered",
                "awaiting_deliver": "ordered",
                "delivering": "delivering",
                # FBS-only статусы
                "acceptance_in_progress": "ordered",
                "awaiting_registration": "ordered",
                "not_accepted": "cancelled",
                "arbitration": "returned",
                "client_arbitration": "returned",
                "awaiting_approve": "ordered",
                "awaiting_verification": "ordered",
                "driver_pickup": "delivering",
                "sent_by_seller": "delivering",
                "cancelled_from_split_pending": "cancelled",
            }

            batch = []
            for posting in all_postings:
                posting_number = posting.get("posting_number", "")
                if not posting_number:
                    continue

                ozon_status = posting.get("status", "")
                our_status = STATUS_MAP.get(ozon_status, "ordered")

                analytics = posting.get("analytics_data") or {}
                financial = posting.get("financial_data") or {}

                products_list = posting.get("products") or []
                in_process_at = posting.get("in_process_at") or posting.get("created_at") or started_at.isoformat()
                shipment_date = posting.get("shipment_date")

                for idx, product in enumerate(products_list):
                    sku = str(product.get("sku", ""))
                    offer_id = product.get("offer_id", "")
                    quantity = product.get("quantity", 1)
                    price_val = float(product.get("price", "0") or "0")

                    # Определяем barcode и product_id
                    barcode = offer_id  # offer_id в Ozon = barcode
                    product_id = None
                    if barcode and barcode in products_map:
                        product_id = products_map[barcode]["id"]
                    elif sku and sku in ozon_sku_map:
                        barcode = ozon_sku_map[sku]
                        if barcode in products_map:
                            product_id = products_map[barcode]["id"]

                    # order_id: для мульти-товарных — суффикс
                    order_id = posting_number
                    if len(products_list) > 1:
                        order_id = f"{posting_number}_{sku}"

                    # Финансовые данные — два формата:
                    # FBO: financial_data.products[] (per-product commission/payout)
                    # FBS: financial_data.commission_amount, financial_data.payout (top-level)
                    commission_amount = 0.0
                    payout_val = None

                    fin_products = financial.get("products", [])
                    if fin_products:
                        # FBO формат: ищем финансы по product_id или по индексу
                        product_sku_int = int(sku) if sku.isdigit() else 0
                        fin_item = None
                        for fp in fin_products:
                            if fp.get("product_id") == product_sku_int:
                                fin_item = fp
                                break
                        if not fin_item and idx < len(fin_products):
                            fin_item = fin_products[idx]

                        if fin_item:
                            commission_amount = abs(float(fin_item.get("commission_amount", 0) or 0))
                            payout_val = float(fin_item.get("payout", 0) or 0)
                            # FBO price может быть точнее (с учётом скидок)
                            fin_price = float(fin_item.get("price", 0) or 0)
                            if fin_price and fin_price > 0:
                                price_val = fin_price
                    else:
                        # FBS формат: top-level fields
                        product_count = len(products_list)
                        commission_amount = abs(float(financial.get("commission_amount", 0) or 0)) / product_count
                        payout_val = float(financial.get("payout", 0) or 0) / product_count if financial.get("payout") else None

                    # --- Парсинг posting_services + item_services (логистика, фулфилмент, возвраты) ---
                    # FBS: posting_services заполнены, item_services = deprecated/нули
                    # FBO: posting_services = нули, item_services заполнены (per-product)
                    # Стратегия: posting_services делятся на кол-во товаров + item_services конкретного товара
                    posting_services = financial.get("posting_services", {}) or {}
                    product_count = len(products_list)
                    # item_services per-product (НЕ суммарные по всем products)
                    product_item_services = {}
                    if fin_products and idx < len(fin_products):
                        for svc_key, svc_val in (fin_products[idx].get("item_services", {}) or {}).items():
                            product_item_services[svc_key] = abs(float(svc_val or 0))

                    def _get_svc(key):
                        # posting_services — общие на весь posting, делим на product_count
                        ps_val = abs(float(posting_services.get(key, 0) or 0)) / product_count
                        return ps_val + product_item_services.get(key, 0)

                    logistics_cost = sum([
                        _get_svc("marketplace_service_item_deliv_to_customer"),
                        _get_svc("marketplace_service_item_direct_flow_trans"),
                        _get_svc("marketplace_service_item_return_flow_trans"),
                        _get_svc("marketplace_service_item_pickup"),
                    ])
                    fulfillment_cost = sum([
                        _get_svc("marketplace_service_item_dropoff_ff"),
                        _get_svc("marketplace_service_item_dropoff_pvz"),
                        _get_svc("marketplace_service_item_dropoff_sc"),
                        _get_svc("marketplace_service_item_fulfillment"),
                    ])
                    return_cost = sum([
                        _get_svc("marketplace_service_item_return_after_deliv_to_customer"),
                        _get_svc("marketplace_service_item_return_not_deliv_to_customer"),
                        _get_svc("marketplace_service_item_return_part_goods_customer"),
                    ])
                    other_cost = fulfillment_cost + return_cost

                    settled = our_status == "sold" and payout_val is not None and payout_val > 0

                    fulfillment = posting.get("_fulfillment", "FBO")
                    total_price = price_val * quantity
                    row = {
                        "marketplace": "ozon",
                        "order_id": order_id,
                        "product_id": product_id,
                        "barcode": barcode,
                        "order_date": in_process_at,
                        "last_change_date": shipment_date,
                        "status": our_status,
                        "price": total_price,
                        "sale_price": total_price,  # Ozon: цена = реальная цена (нет скрытой СПП)
                        "sale_amount": payout_val,
                        "commission": round(commission_amount, 2),
                        "logistics": round(logistics_cost, 2),
                        "storage_fee": 0,
                        "other_fees": round(other_cost, 2),
                        "payout": payout_val,
                        "settled": settled,
                        "quantity": quantity,
                        "ozon_posting_status": f"{fulfillment}:{ozon_status}",
                        # FBO: нет region (fallback на city), warehouse_name
                        # FBS: region доступен, warehouse (не warehouse_name)
                        "region": (
                            analytics.get("city", "")
                            if fulfillment == "FBO"
                            else analytics.get("region", "") or analytics.get("city", "")
                        ),
                        "warehouse": (
                            analytics.get("warehouse_name", "")
                            if fulfillment == "FBO"
                            else analytics.get("warehouse", "") or analytics.get("warehouse_name", "")
                        ),
                        "updated_at": datetime.now().isoformat(),
                        "fulfillment_type": fulfillment,
                    }

                    # Причины отмен (Phase 5)
                    if fulfillment == "FBS":
                        cancellation = posting.get("cancellation", {}) or {}
                        row["cancel_reason"] = (cancellation.get("cancel_reason", "") or "")[:200]
                        row["cancellation_initiator"] = (cancellation.get("cancellation_initiator", "") or "")[:20]
                    else:
                        # FBO: только cancel_reason_id (число), нет cancellation объекта
                        cancel_id = posting.get("cancel_reason_id")
                        row["cancel_reason"] = str(cancel_id)[:200] if cancel_id else ""
                        row["cancellation_initiator"] = ""

                    if self.user_id:
                        row["user_id"] = self.user_id
                    batch.append(row)

            # Upsert пачками по 500
            for i in range(0, len(batch), 500):
                chunk = batch[i:i+500]
                self.supabase.table("mp_orders").upsert(
                    chunk, on_conflict="user_id,marketplace,order_id"
                ).execute()
                records_count += len(chunk)

            self._log_sync("ozon", "orders", "success", records_count, started_at=started_at)
            return {"status": "success", "records": records_count}

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Ошибка синхронизации заказов Ozon: {error_msg}")
            self._log_sync("ozon", "orders", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    # ==================== ХРАНЕНИЕ OZON (per-product) ====================

    def _build_legacy_storage_rows(
        self,
        daily_agg: dict,
        ozon_sku_map: dict,
        products_map: dict,
        date_from_str: str,
        date_to_str: str,
    ) -> list[dict]:
        """
        Aggregate daily_agg back to period-level for legacy mp_storage_costs table.
        Keeps backward compatibility while we transition to mp_storage_costs_daily.
        """
        # Re-aggregate by SKU (sum cost, max qty/volume across all dates)
        sku_totals: dict[str, dict] = {}  # {sku_str: {cost, qty, volume_ml, barcode}}
        for (_, sku_str), agg in daily_agg.items():
            if sku_str not in sku_totals:
                sku_totals[sku_str] = {"cost": 0.0, "qty": 0, "volume_ml": 0.0, "barcode": agg["barcode"]}
            sku_totals[sku_str]["cost"] += agg["cost"]
            sku_totals[sku_str]["qty"] = max(sku_totals[sku_str]["qty"], agg["qty"])
            sku_totals[sku_str]["volume_ml"] = max(sku_totals[sku_str]["volume_ml"], agg["volume_ml"])

        rows = []
        for sku_str, totals in sku_totals.items():
            if totals["cost"] <= 0:
                continue
            barcode = ozon_sku_map.get(sku_str) or totals["barcode"]
            product = products_map.get(barcode)
            if not product:
                continue
            row = {
                "marketplace": "ozon",
                "product_id": product["id"],
                "date_from": date_from_str,
                "date_to": date_to_str,
                "storage_cost": round(totals["cost"], 2),
                "quantity": totals["qty"],
                "volume_liters": round(totals["volume_ml"] / 1000, 2),
            }
            if self.user_id:
                row["user_id"] = self.user_id
            rows.append(row)
        return rows

    async def sync_storage_ozon(self, date_from: datetime, date_to: datetime) -> dict:
        """
        Синхронизация per-product storage costs из Ozon Placement Report API.
        API: /v1/report/placement/by-products/create → /v1/report/info → download XLSX.
        Лимит: 5 вызовов в день. Макс. период: 31 день.
        """
        import asyncio
        import io
        import httpx as _httpx
        from openpyxl import load_workbook

        started_at = datetime.now()
        records_count = 0
        date_from_str = date_from.strftime("%Y-%m-%d")
        date_to_str = date_to.strftime("%Y-%m-%d")

        try:
            # 1. Создаём отчёт
            logger.info(f"Ozon storage: creating placement report {date_from_str} - {date_to_str}")
            code = await self.ozon_client.create_placement_report(date_from_str, date_to_str)
            if not code:
                raise RuntimeError("Ozon placement report: empty code returned")

            # 2. Поллим статус (каждые 5 сек, макс 2 мин = 24 попытки)
            file_url = None
            for attempt in range(24):
                await asyncio.sleep(5)
                info = await self.ozon_client.get_report_info(code)
                status = info.get("status", "")
                logger.info(f"Ozon storage report status: {status} (attempt {attempt + 1})")

                if status == "success":
                    file_url = info.get("file", "")
                    break
                elif status == "failed":
                    error_msg = info.get("error", "unknown error")
                    raise RuntimeError(f"Ozon placement report failed: {error_msg}")
                # waiting / processing — continue polling

            if not file_url:
                raise RuntimeError("Ozon placement report: timeout (2 min), report not ready")

            # 3. Скачиваем XLSX
            logger.info(f"Ozon storage: downloading XLSX from {file_url[:80]}...")
            async with _httpx.AsyncClient(timeout=60.0, follow_redirects=True) as http_client:
                response = await http_client.get(file_url)
                response.raise_for_status()
                xlsx_bytes = response.content

            # 4. Парсим XLSX
            wb = load_workbook(filename=io.BytesIO(xlsx_bytes), read_only=True)
            ws = wb.active
            rows_data = list(ws.iter_rows(values_only=True))
            wb.close()

            if len(rows_data) < 3:
                logger.warning("Ozon storage: XLSX has less than 3 rows (no data)")
                self._log_sync("ozon", "storage", "success", 0, started_at=started_at)
                return {"status": "success", "records": 0}

            # XLSX structure (from Ozon Placement Report API):
            # Row 0 = headers: Дата | SKU | Артикул | Категория | Тип | Склад | Признак |
            #   Суммарный объем(мл) | Кол-во экз | Платный объем(мл) | Кол-во платных | Начисленная стоимость
            # Row 1+ = data (daily, per-warehouse → aggregate by date+SKU)
            # Col[0] = date, Col[1] = SKU (int), Col[2] = barcode, Col[7] = volume(ml),
            # Col[8] = qty, Col[11] = cost(RUB)
            ozon_sku_map = self.ozon_sku_map  # {str(sku): barcode}
            products_map = self._get_products_map()  # {barcode: {id, ...}}

            # Aggregate per (date, SKU): sum cost across warehouses, max qty/volume per day
            # Key = (date_str, sku_str)
            daily_agg: dict[tuple[str, str], dict] = {}
            for row in rows_data[1:]:  # skip header row
                if not row or len(row) < 12:
                    continue

                # Parse date (col 0) — may be datetime object or string
                date_val = row[0]
                if date_val is None:
                    continue
                if isinstance(date_val, datetime):
                    row_date_str = date_val.strftime("%Y-%m-%d")
                elif hasattr(date_val, "strftime"):
                    row_date_str = date_val.strftime("%Y-%m-%d")
                else:
                    # String like "2026-02-21" or "21.02.2026"
                    raw = str(date_val).strip()
                    if "." in raw and len(raw) == 10:
                        # DD.MM.YYYY → YYYY-MM-DD
                        parts = raw.split(".")
                        if len(parts) == 3:
                            row_date_str = f"{parts[2]}-{parts[1]}-{parts[0]}"
                        else:
                            continue
                    else:
                        row_date_str = raw[:10]  # "YYYY-MM-DD..."

                sku_val = row[1]
                if not sku_val:
                    continue
                sku_str = str(int(sku_val)) if isinstance(sku_val, (float, int)) else str(sku_val).strip()

                # Parse storage cost (col 11)
                cost_val = row[11]
                if cost_val is None:
                    continue
                try:
                    cost = float(cost_val) if isinstance(cost_val, (int, float)) else float(str(cost_val).replace(",", ".").replace("\xa0", "").strip())
                except (ValueError, TypeError):
                    continue

                # Parse quantity (col 8) and volume (col 7)
                qty = 0
                vol_ml = 0.0
                try:
                    if row[8] is not None:
                        qty = int(float(row[8])) if isinstance(row[8], (int, float)) else int(float(str(row[8]).strip()))
                except (ValueError, TypeError):
                    pass
                try:
                    if row[7] is not None:
                        vol_ml = float(row[7]) if isinstance(row[7], (int, float)) else float(str(row[7]).strip())
                except (ValueError, TypeError):
                    pass

                key = (row_date_str, sku_str)
                if key not in daily_agg:
                    daily_agg[key] = {"cost": 0.0, "qty": 0, "volume_ml": 0.0, "barcode": str(row[2] or "")}
                daily_agg[key]["cost"] += cost
                daily_agg[key]["qty"] = max(daily_agg[key]["qty"], qty)
                daily_agg[key]["volume_ml"] = max(daily_agg[key]["volume_ml"], vol_ml)

            unique_skus = len(set(k[1] for k in daily_agg))
            unique_dates = len(set(k[0] for k in daily_agg))
            logger.info(f"Ozon storage: aggregated {len(daily_agg)} (date,SKU) pairs "
                        f"({unique_skus} SKUs, {unique_dates} dates) from {len(rows_data)-1} rows")

            # Build insert rows for mp_storage_costs_daily
            insert_rows = []
            for (row_date, sku_str), agg in daily_agg.items():
                storage_cost = agg["cost"]
                if storage_cost <= 0:
                    continue

                # Map SKU → product_id (try ozon_sku_map first, then barcode direct)
                barcode = ozon_sku_map.get(sku_str) or agg["barcode"]
                product = products_map.get(barcode)
                if not product:
                    logger.debug(f"Ozon storage: SKU {sku_str} / barcode {barcode} not in products_map, skipping")
                    continue

                product_id = product["id"]
                volume_liters = round(agg["volume_ml"] / 1000, 2)

                insert_row = {
                    "marketplace": "ozon",
                    "product_id": product_id,
                    "date": row_date,
                    "storage_cost": round(storage_cost, 2),
                    "quantity": agg["qty"],
                    "volume_liters": volume_liters,
                }
                if self.user_id:
                    insert_row["user_id"] = self.user_id

                insert_rows.append(insert_row)

            if insert_rows:
                logger.info(f"Ozon storage: inserting {len(insert_rows)} daily rows "
                            f"for {date_from_str} - {date_to_str}")

            # Also write to legacy mp_storage_costs (period-based, for backward compat)
            legacy_rows = self._build_legacy_storage_rows(daily_agg, ozon_sku_map, products_map,
                                                          date_from_str, date_to_str)

            # 5. UPSERT daily rows (delete period + insert)
            if insert_rows:
                # Delete existing daily records for this date range
                del_query = self.supabase.table("mp_storage_costs_daily") \
                    .delete() \
                    .eq("marketplace", "ozon") \
                    .gte("date", date_from_str) \
                    .lte("date", date_to_str)
                if self.user_id:
                    del_query = del_query.eq("user_id", self.user_id)
                del_query.execute()

                # Insert daily rows in batches (Supabase limit)
                batch_size = 500
                for i in range(0, len(insert_rows), batch_size):
                    batch = insert_rows[i:i + batch_size]
                    self.supabase.table("mp_storage_costs_daily").insert(batch).execute()
                records_count = len(insert_rows)

            # 5b. Also upsert legacy mp_storage_costs (period-based)
            if legacy_rows:
                del_legacy = self.supabase.table("mp_storage_costs") \
                    .delete() \
                    .eq("marketplace", "ozon") \
                    .eq("date_from", date_from_str) \
                    .eq("date_to", date_to_str)
                if self.user_id:
                    del_legacy = del_legacy.eq("user_id", self.user_id)
                del_legacy.execute()
                self.supabase.table("mp_storage_costs").insert(legacy_rows).execute()

            logger.info(f"Ozon storage sync: {records_count} products saved for {date_from_str} - {date_to_str}")
            self._log_sync("ozon", "storage", "success", records_count, started_at=started_at)
            return {"status": "success", "records": records_count}

        except Exception as e:
            import traceback
            error_msg = str(e) or repr(e)
            tb = traceback.format_exc()
            logger.error(f"Ошибка синхронизации storage Ozon: {error_msg}\n{tb}")
            self._log_sync("ozon", "storage", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    # ==================== WB PAID STORAGE ====================

    async def sync_storage_wb(self, date_from: datetime, date_to: datetime) -> dict:
        """
        Синхронизация per-product storage costs из WB Paid Storage API.
        API: GET /api/v1/paid_storage (3-step async: create task → poll → download).
        Макс. период: 8 дней (wb_client handles chunking).
        Результат: UPSERT в mp_storage_costs_daily с marketplace='wb'.
        """
        started_at = datetime.now()
        records_count = 0
        date_from_str = date_from.strftime("%Y-%m-%d")
        date_to_str = date_to.strftime("%Y-%m-%d")

        try:
            if not self.wb_client:
                logger.info("WB storage sync skipped: no WB token configured")
                return {"status": "success", "records": 0}

            # 1. Fetch paid storage data from WB API
            logger.info(f"WB storage: fetching paid_storage {date_from_str} - {date_to_str}")
            raw_rows = await self.wb_client.get_paid_storage(date_from_str, date_to_str)

            if not raw_rows:
                logger.info("WB storage: no data returned from API")
                self._log_sync("wb", "storage", "success", 0, started_at=started_at)
                return {"status": "success", "records": 0}

            logger.info(f"WB storage: received {len(raw_rows)} raw rows from API")

            # 2. Build nmId → product_id map
            products_map = self._get_products_map()  # {barcode: product_dict}
            nm_id_to_product: dict[int, dict] = {}
            barcode_to_product: dict[str, dict] = {}
            for barcode, product in products_map.items():
                if barcode and barcode != "WB_ACCOUNT":
                    barcode_to_product[barcode] = product
                    wb_nm_id = product.get("wb_nm_id")
                    if wb_nm_id:
                        nm_id_to_product[int(wb_nm_id)] = product

            # 3. Aggregate: per (date, product_id) → sum warehousePrice across warehouses/calcTypes
            # Key = (date_str, product_id)
            daily_agg: dict[tuple[str, str], dict] = {}

            for row in raw_rows:
                row_date = row.get("date") or row.get("originalDate")
                if not row_date:
                    continue
                row_date_str = str(row_date)[:10]  # "YYYY-MM-DD"

                # Map to product: nmId first, barcode fallback
                nm_id = row.get("nmId")
                barcode = row.get("barcode", "")
                product = None
                if nm_id:
                    product = nm_id_to_product.get(int(nm_id))
                if not product and barcode:
                    product = barcode_to_product.get(barcode)

                if not product:
                    continue  # Unknown product, skip

                product_id = product["id"]
                warehouse_price = float(row.get("warehousePrice", 0) or 0)
                barcodes_count = int(row.get("barcodesCount", 0) or 0)
                volume = float(row.get("volume", 0) or 0)

                key = (row_date_str, product_id)
                if key not in daily_agg:
                    daily_agg[key] = {"cost": 0.0, "qty": 0, "volume": 0.0}

                # warehousePrice can be negative (discounts like "скидка на остаток склада")
                # We SUM all calcTypes: base charge + discounts = net cost per product per day
                daily_agg[key]["cost"] += warehouse_price
                daily_agg[key]["qty"] = max(daily_agg[key]["qty"], barcodes_count)
                daily_agg[key]["volume"] = max(daily_agg[key]["volume"], volume)

            unique_products = len(set(k[1] for k in daily_agg))
            unique_dates = len(set(k[0] for k in daily_agg))
            logger.info(f"WB storage: aggregated {len(daily_agg)} (date,product) pairs "
                        f"({unique_products} products, {unique_dates} dates) from {len(raw_rows)} rows")

            # 4. Build insert rows
            insert_rows = []
            for (row_date, product_id), agg in daily_agg.items():
                storage_cost = agg["cost"]
                # Skip zero/negative net storage (fully discounted products)
                if storage_cost <= 0:
                    continue

                insert_row = {
                    "marketplace": "wb",
                    "product_id": product_id,
                    "date": row_date,
                    "storage_cost": round(storage_cost, 2),
                    "quantity": agg["qty"],
                    "volume_liters": round(agg["volume"], 2),
                }
                if self.user_id:
                    insert_row["user_id"] = self.user_id
                insert_rows.append(insert_row)

            # 5. UPSERT: delete period + insert
            if insert_rows:
                logger.info(f"WB storage: inserting {len(insert_rows)} daily rows "
                            f"for {date_from_str} - {date_to_str}")
                del_query = self.supabase.table("mp_storage_costs_daily") \
                    .delete() \
                    .eq("marketplace", "wb") \
                    .gte("date", date_from_str) \
                    .lte("date", date_to_str)
                if self.user_id:
                    del_query = del_query.eq("user_id", self.user_id)
                del_query.execute()

                # Insert in batches
                batch_size = 500
                for i in range(0, len(insert_rows), batch_size):
                    batch = insert_rows[i:i + batch_size]
                    self.supabase.table("mp_storage_costs_daily").insert(batch).execute()
                records_count = len(insert_rows)

            logger.info(f"WB storage sync: {records_count} rows saved for {date_from_str} - {date_to_str}")
            self._log_sync("wb", "storage", "success", records_count, started_at=started_at)
            return {"status": "success", "records": records_count}

        except Exception as e:
            import traceback
            error_msg = str(e) or repr(e)
            tb = traceback.format_exc()
            logger.error(f"Ошибка синхронизации storage WB: {error_msg}\n{tb}")
            self._log_sync("wb", "storage", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    # ==================== DELIVERY DATES OZON ====================

    async def sync_delivery_dates_ozon(self, date_from: datetime, date_to: datetime) -> dict:
        """
        Синхронизация delivery_date из Ozon Postings Report CSV.
        API: /v1/report/postings/create → /v1/report/info → download CSV.
        Два запроса: FBO + FBS (delivery_schema принимает одно значение за раз).
        Результат: UPDATE mp_orders SET delivery_date WHERE order_id = posting_number.
        """
        import asyncio
        import csv
        import io
        import httpx as _httpx

        started_at = datetime.now()
        updated_count = 0
        date_from_iso = date_from.strftime("%Y-%m-%dT00:00:00.000Z")
        date_to_iso = date_to.strftime("%Y-%m-%dT23:59:59.999Z")

        try:
            all_rows: list[dict] = []

            for schema in ("fbo", "fbs"):
                # 1. Создаём отчёт
                logger.info(f"Ozon delivery dates: creating postings report ({schema}) "
                            f"{date_from.strftime('%Y-%m-%d')} - {date_to.strftime('%Y-%m-%d')}")
                try:
                    code = await self.ozon_client.create_postings_report(
                        date_from_iso, date_to_iso, delivery_schema=schema
                    )
                except Exception as e:
                    logger.warning(f"Ozon delivery dates: failed to create {schema} report: {e}")
                    continue

                if not code:
                    logger.warning(f"Ozon delivery dates: empty code for {schema} report")
                    continue

                # 2. Поллим статус (каждые 5 сек, макс 2 мин = 24 попытки)
                file_url = None
                for attempt in range(24):
                    await asyncio.sleep(5)
                    info = await self.ozon_client.get_report_info(code)
                    status = info.get("status", "")
                    logger.info(f"Ozon delivery report ({schema}) status: {status} (attempt {attempt + 1})")

                    if status == "success":
                        file_url = info.get("file", "")
                        break
                    elif status == "failed":
                        error_msg = info.get("error", "unknown error")
                        logger.warning(f"Ozon delivery report ({schema}) failed: {error_msg}")
                        break
                    # waiting / processing — continue polling

                if not file_url:
                    logger.warning(f"Ozon delivery report ({schema}): no file URL (timeout or failed)")
                    continue

                # 3. Скачиваем CSV
                logger.info(f"Ozon delivery dates: downloading CSV ({schema}) from {file_url[:80]}...")
                async with _httpx.AsyncClient(timeout=60.0, follow_redirects=True) as http_client:
                    response = await http_client.get(file_url)
                    response.raise_for_status()
                    csv_bytes = response.content

                # 4. Парсим CSV (delimiter=';', UTF-8 with BOM)
                csv_text = csv_bytes.decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(csv_text), delimiter=";")

                schema_rows = 0
                for row in reader:
                    posting_number = (row.get("Номер отправления") or "").strip()
                    delivery_date_str = (row.get("Дата доставки") or "").strip()
                    status_val = (row.get("Статус") or "").strip()

                    # Only delivered orders have a delivery date
                    if not posting_number or not delivery_date_str or status_val != "Доставлен":
                        continue

                    # Parse delivery_date: "2026-02-24 12:51:08" (Ozon CSV = Moscow TZ, Rule #42)
                    from datetime import timezone as _tz
                    _MSK = _tz(timedelta(hours=3))
                    try:
                        delivery_dt = datetime.strptime(delivery_date_str, "%Y-%m-%d %H:%M:%S").replace(tzinfo=_MSK)
                    except ValueError:
                        try:
                            delivery_dt = datetime.strptime(delivery_date_str[:10], "%Y-%m-%d").replace(tzinfo=_MSK)
                        except ValueError:
                            logger.debug(f"Ozon delivery dates: unparseable date '{delivery_date_str}' for {posting_number}")
                            continue

                    sku = (row.get("SKU") or "").strip()
                    quantity_str = (row.get("Количество") or "1").strip()
                    try:
                        quantity = int(quantity_str)
                    except ValueError:
                        quantity = 1

                    all_rows.append({
                        "posting_number": posting_number,
                        "delivery_date": delivery_dt.isoformat(),
                        "sku": sku,
                        "quantity": quantity,
                    })
                    schema_rows += 1

                logger.info(f"Ozon delivery dates ({schema}): parsed {schema_rows} delivered rows from CSV")

            if not all_rows:
                logger.info("Ozon delivery dates: no delivered rows found in CSV reports")
                self._log_sync("ozon", "delivery_dates", "success", 0, started_at=started_at)
                return {"status": "success", "records": 0}

            # 5. Batch UPDATE mp_orders via RPC (Rule #49: PERF-002 fix)
            # Replaces N+1 individual UPDATEs with single SQL call via UNNEST.
            # mp_orders.order_id = posting_number (single) or posting_number_sku (multi).

            # Group by posting_number to detect multi-item postings
            postings_by_number: dict[str, list[dict]] = {}
            for row in all_rows:
                pn = row["posting_number"]
                if pn not in postings_by_number:
                    postings_by_number[pn] = []
                postings_by_number[pn].append(row)

            # Build batch arrays for RPC
            batch_order_ids: list[str] = []
            batch_delivery_dates: list[str] = []
            batch_quantities: list[int] = []

            for posting_number, rows in postings_by_number.items():
                delivery_date_iso = rows[0]["delivery_date"]
                is_multi_item = len(rows) > 1

                if is_multi_item:
                    for row in rows:
                        order_id = f"{posting_number}_{row['sku']}"
                        batch_order_ids.append(order_id)
                        batch_delivery_dates.append(delivery_date_iso)
                        batch_quantities.append(row.get("quantity", 1))
                else:
                    # Single-item: try both posting_number and posting_number_sku
                    batch_order_ids.append(posting_number)
                    batch_delivery_dates.append(delivery_date_iso)
                    batch_quantities.append(rows[0].get("quantity", 1))
                    # Also add with SKU suffix as fallback
                    sku = rows[0].get("sku", "")
                    if sku:
                        batch_order_ids.append(f"{posting_number}_{sku}")
                        batch_delivery_dates.append(delivery_date_iso)
                        batch_quantities.append(rows[0].get("quantity", 1))

            # Execute single batch RPC call (replaces ~600+ individual HTTP requests)
            if batch_order_ids and self.user_id:
                try:
                    result = self.supabase.rpc("batch_update_delivery_dates", {
                        "p_user_id": self.user_id,
                        "p_order_ids": batch_order_ids,
                        "p_delivery_dates": batch_delivery_dates,
                        "p_quantities": batch_quantities,
                    }).execute()
                    updated_count = result.data if isinstance(result.data, int) else 0
                except Exception as e:
                    logger.warning(f"Ozon delivery dates: batch RPC failed, falling back to individual updates: {e}")
                    # Fallback: individual updates (for backward compat if RPC not yet deployed)
                    for i, order_id in enumerate(batch_order_ids):
                        try:
                            self.supabase.table("mp_orders").update({
                                "delivery_date": batch_delivery_dates[i],
                                "quantity": batch_quantities[i],
                            }).eq("user_id", self.user_id).eq("marketplace", "ozon").eq("order_id", order_id).execute()
                            updated_count += 1
                        except Exception:
                            pass

            logger.info(f"Ozon delivery dates sync: {updated_count} orders updated with delivery_date "
                        f"(from {len(all_rows)} CSV rows, {len(postings_by_number)} unique postings)")
            self._log_sync("ozon", "delivery_dates", "success", updated_count, started_at=started_at)
            return {"status": "success", "records": updated_count}

        except Exception as e:
            import traceback
            error_msg = str(e) or repr(e)
            tb = traceback.format_exc()
            logger.error(f"Ошибка синхронизации delivery dates Ozon: {error_msg}\n{tb}")
            self._log_sync("ozon", "delivery_dates", "error", 0, error_msg, started_at)
            return {"status": "error", "message": error_msg}

    # ==================== ПОЛНАЯ СИНХРОНИЗАЦИЯ ====================

    async def sync_all(self, days_back: int = 30) -> dict:
        """
        Полная синхронизация всех данных
        days_back - за сколько дней загружать данные
        """
        date_from = datetime.now() - timedelta(days=days_back)
        date_to = datetime.now()

        results = {
            "products": await self.sync_products(),
            "orders_wb": await self.sync_orders_wb(date_from, date_to),
            "orders_ozon": await self.sync_orders_ozon(date_from, date_to),
            "sales_wb": await self.sync_sales_wb(date_from, date_to),
            "sales_ozon": await self.sync_sales_ozon(date_from, date_to),
            "stocks_wb": await self.sync_stocks_wb(),
            "stocks_ozon": await self.sync_stocks_ozon(),
            "costs_wb": await self.sync_costs_wb(date_from, date_to),
            "costs_ozon": await self.sync_costs_ozon(date_from, date_to),
            "ads_wb": await self.sync_ads_wb(date_from, date_to),
            "ads_ozon": await self.sync_ads_ozon(date_from, date_to),
        }

        # Подсчёт успешных/ошибок
        success = sum(1 for r in results.values() if r.get("status") == "success")
        errors = sum(1 for r in results.values() if r.get("status") == "error")

        return {
            "status": "completed",
            "success_count": success,
            "error_count": errors,
            "details": results
        }
