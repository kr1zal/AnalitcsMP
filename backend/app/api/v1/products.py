"""
Роутер для работы с товарами
"""
import csv
import io
import logging
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from pydantic import BaseModel, Field
from typing import Optional

from ...db.supabase import get_supabase_client
from ...auth import CurrentUser, get_current_user
from ...subscription import _load_subscription

logger = logging.getLogger(__name__)

router = APIRouter()


# ==================== PYDANTIC MODELS ====================

class CreateProductRequest(BaseModel):
    name: str = Field(..., min_length=1, max_length=255)
    barcode: str = Field(..., min_length=1, max_length=100)
    purchase_price: float = Field(..., ge=0)
    wb_nm_id: Optional[int] = None
    wb_vendor_code: Optional[str] = None
    ozon_product_id: Optional[int] = None
    ozon_offer_id: Optional[str] = None


class UpdatePurchasePriceRequest(BaseModel):
    purchase_price: float = Field(..., ge=0)


class ReorderItem(BaseModel):
    product_id: str
    sort_order: int


class ReorderRequest(BaseModel):
    items: list[ReorderItem]


class LinkProductsRequest(BaseModel):
    wb_product_id: str
    ozon_product_id: str
    purchase_price: float = Field(..., ge=0)


# ==================== GET ENDPOINTS ====================

@router.get("/products")
async def get_products(
    current_user: CurrentUser = Depends(get_current_user),
    marketplace: Optional[str] = None,
):
    """
    Получить список всех товаров (исключая WB_ACCOUNT системный товар)
    """
    supabase = get_supabase_client()

    try:
        query = (
            supabase.table("mp_products")
            .select("*")
            .eq("user_id", current_user.id)
            .neq("barcode", "WB_ACCOUNT")
            .order("sort_order")
            .limit(500)
        )

        if marketplace == "wb":
            query = query.not_.is_("wb_nm_id", "null")
        elif marketplace == "ozon":
            query = query.not_.is_("ozon_product_id", "null")

        result = query.execute()

        return {
            "status": "success",
            "count": len(result.data),
            "products": result.data
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/products/{product_id}")
async def get_product_by_id(
    product_id: str,
    current_user: CurrentUser = Depends(get_current_user),
):
    """
    Получить товар по ID (UUID)
    """
    supabase = get_supabase_client()

    try:
        result = supabase.table("mp_products").select("*").eq("id", product_id).eq("user_id", current_user.id).execute()

        if not result.data:
            raise HTTPException(status_code=404, detail="Product not found")

        return {
            "status": "success",
            "product": result.data[0]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/products/barcode/{barcode}")
async def get_product_by_barcode(
    barcode: str,
    current_user: CurrentUser = Depends(get_current_user),
):
    """
    Получить товар по штрихкоду
    """
    supabase = get_supabase_client()

    try:
        result = supabase.table("mp_products").select("*").eq("barcode", barcode).eq("user_id", current_user.id).execute()

        if not result.data:
            raise HTTPException(status_code=404, detail="Product not found")

        return {
            "status": "success",
            "product": result.data[0]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== MUTATION ENDPOINTS ====================

@router.post("/products")
async def create_product(
    body: CreateProductRequest,
    current_user: CurrentUser = Depends(get_current_user),
):
    """
    Создать товар вручную.
    Проверяет лимит SKU по тарифу перед созданием.
    """
    supabase = get_supabase_client()

    # Проверка лимита SKU по тарифу
    sub = _load_subscription(current_user.id)
    max_sku = sub.plan_config.get("max_sku")

    if max_sku is not None:
        products = (
            supabase.table("mp_products")
            .select("id")
            .eq("user_id", current_user.id)
            .neq("barcode", "WB_ACCOUNT")
            .execute()
        )
        current_count = len(products.data) if products.data else 0
        if current_count >= max_sku:
            raise HTTPException(
                status_code=403,
                detail={
                    "error": "sku_limit_reached",
                    "message": f"Лимит товаров ({max_sku}) достигнут. Перейдите на Pro для увеличения.",
                    "current_plan": sub.plan,
                    "max_sku": max_sku,
                    "current_sku": current_count,
                },
            )

    try:
        # Проверить уникальность штрихкода для этого пользователя
        existing = (
            supabase.table("mp_products")
            .select("id")
            .eq("barcode", body.barcode)
            .eq("user_id", current_user.id)
            .execute()
        )
        if existing.data:
            raise HTTPException(
                status_code=409,
                detail=f"Товар с штрихкодом '{body.barcode}' уже существует",
            )

        now = datetime.now(timezone.utc).isoformat()
        row: dict = {
            "user_id": current_user.id,
            "barcode": body.barcode,
            "name": body.name,
            "purchase_price": body.purchase_price,
            "updated_at": now,
        }
        if body.wb_nm_id is not None:
            row["wb_nm_id"] = body.wb_nm_id
        if body.wb_vendor_code is not None:
            row["wb_vendor_code"] = body.wb_vendor_code
        if body.ozon_product_id is not None:
            row["ozon_product_id"] = body.ozon_product_id
        if body.ozon_offer_id is not None:
            row["ozon_offer_id"] = body.ozon_offer_id

        result = supabase.table("mp_products").insert(row).execute()

        if not result.data:
            raise HTTPException(status_code=500, detail="Не удалось создать товар")

        return {
            "status": "success",
            "product": result.data[0],
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/products/{product_id}/purchase-price")
async def update_purchase_price(
    product_id: str,
    body: UpdatePurchasePriceRequest,
    current_user: CurrentUser = Depends(get_current_user),
):
    """
    Обновить себестоимость товара.
    Если товар в группе (product_group_id) — обновляет все связанные товары.
    """
    supabase = get_supabase_client()

    try:
        # Найти товар
        result = (
            supabase.table("mp_products")
            .select("id, product_group_id")
            .eq("id", product_id)
            .eq("user_id", current_user.id)
            .execute()
        )
        if not result.data:
            raise HTTPException(status_code=404, detail="Product not found")

        product = result.data[0]
        now = datetime.now(timezone.utc).isoformat()

        # Обновить сам товар
        supabase.table("mp_products").update({
            "purchase_price": body.purchase_price,
            "updated_at": now,
        }).eq("id", product_id).eq("user_id", current_user.id).execute()

        linked_updated = 0

        # Если товар в группе — обновить остальных в группе
        group_id = product.get("product_group_id")
        if group_id:
            linked_result = (
                supabase.table("mp_products")
                .update({"purchase_price": body.purchase_price, "updated_at": now})
                .eq("product_group_id", group_id)
                .eq("user_id", current_user.id)
                .neq("id", product_id)
                .execute()
            )
            linked_updated = len(linked_result.data) if linked_result.data else 0

        return {
            "status": "success",
            "product_id": product_id,
            "purchase_price": body.purchase_price,
            "linked_updated": linked_updated,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/products/reorder")
async def reorder_products(
    body: ReorderRequest,
    current_user: CurrentUser = Depends(get_current_user),
):
    """
    Массовое обновление sort_order для товаров.
    """
    supabase = get_supabase_client()

    try:
        updates = [
            {"id": item.product_id, "sort_order": item.sort_order}
            for item in body.items
        ]
        try:
            result = supabase.rpc("batch_update_products", {
                "p_user_id": current_user.id,
                "p_updates": updates,  # Supabase client сериализует JSONB сам
            }).execute()
            updated = result.data if isinstance(result.data, int) else len(updates)
        except Exception:
            # Fallback: поштучный UPDATE если RPC недоступен
            updated = 0
            now = datetime.now(timezone.utc).isoformat()
            for item in body.items:
                r = supabase.table("mp_products").update({"sort_order": item.sort_order, "updated_at": now}).eq("id", item.product_id).eq("user_id", current_user.id).execute()
                if r.data:
                    updated += 1

        return {"status": "success", "updated": updated}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/products/link")
async def link_products(
    body: LinkProductsRequest,
    current_user: CurrentUser = Depends(get_current_user),
):
    """
    Связать два товара с разных маркетплейсов.
    Устанавливает одинаковый product_group_id и purchase_price на оба.
    """
    supabase = get_supabase_client()

    try:
        # Загрузить оба товара
        wb_result = (
            supabase.table("mp_products")
            .select("id, wb_nm_id, ozon_product_id, product_group_id")
            .eq("id", body.wb_product_id)
            .eq("user_id", current_user.id)
            .execute()
        )
        ozon_result = (
            supabase.table("mp_products")
            .select("id, wb_nm_id, ozon_product_id, product_group_id")
            .eq("id", body.ozon_product_id)
            .eq("user_id", current_user.id)
            .execute()
        )

        if not wb_result.data:
            raise HTTPException(status_code=404, detail="WB product not found")
        if not ozon_result.data:
            raise HTTPException(status_code=404, detail="Ozon product not found")

        wb_product = wb_result.data[0]
        ozon_product = ozon_result.data[0]

        # Валидация: один должен быть WB, другой — Ozon
        if not wb_product.get("wb_nm_id"):
            raise HTTPException(status_code=400, detail="First product must be a WB product (wb_nm_id required)")
        if not ozon_product.get("ozon_product_id"):
            raise HTTPException(status_code=400, detail="Second product must be an Ozon product (ozon_product_id required)")

        # Нельзя связать один и тот же товар
        if wb_product["id"] == ozon_product["id"]:
            raise HTTPException(status_code=400, detail="Cannot link a product with itself")

        # Определить group_id: использовать существующий или новый
        group_id = (
            wb_product.get("product_group_id")
            or ozon_product.get("product_group_id")
            or str(uuid.uuid4())
        )

        now = datetime.now(timezone.utc).isoformat()
        update_data = {
            "product_group_id": group_id,
            "purchase_price": body.purchase_price,
            "updated_at": now,
        }

        # Обновить оба товара
        supabase.table("mp_products").update(update_data).eq("id", body.wb_product_id).eq("user_id", current_user.id).execute()
        supabase.table("mp_products").update(update_data).eq("id", body.ozon_product_id).eq("user_id", current_user.id).execute()

        return {
            "status": "success",
            "group_id": group_id,
            "purchase_price": body.purchase_price,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024  # 5MB
MAX_IMPORT_ROWS = 2000
MAX_PRICE = 999_999

ID_ALIASES = {"barcode", "offer_id", "артикул", "штрихкод", "id", "sku", "баркод"}
PRICE_ALIASES = {"purchase_price", "себестоимость", "цена закупки", "закупка", "цена", "price", "cost"}


def _parse_csv_content(content: bytes) -> list[dict]:
    """Парсит CSV из bytes, возвращает list[{id, price}]."""
    text = content.decode("utf-8-sig").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Файл пуст")

    # Auto-detect delimiter
    first_line = text.split("\n", 1)[0]
    if ";" in first_line:
        delimiter = ";"
    elif "\t" in first_line:
        delimiter = "\t"
    else:
        delimiter = ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows_list = list(reader)
    if len(rows_list) < 2:
        raise HTTPException(status_code=400, detail="Файл должен содержать заголовок и хотя бы одну строку данных")

    headers = [h.strip().strip('"').strip("'").lower() for h in rows_list[0]]

    id_col = next((i for i, h in enumerate(headers) if h in ID_ALIASES), None)
    price_col = next((i for i, h in enumerate(headers) if h in PRICE_ALIASES), None)

    # Fallback: first col = id, last col = price
    if id_col is None and price_col is None and len(headers) >= 2:
        id_col = 0
        price_col = len(headers) - 1

    if id_col is None or price_col is None:
        raise HTTPException(
            status_code=400,
            detail="Не найдены колонки barcode/offer_id и purchase_price. Проверьте заголовки.",
        )

    results = []
    for row in rows_list[1:]:
        if not row or len(row) <= max(id_col, price_col):
            continue
        barcode = row[id_col].strip().strip('"').strip("'")
        if not barcode:
            continue
        price_str = row[price_col].strip().strip('"').strip("'").replace(",", ".")
        # Пустая ячейка = 0 (удаление себестоимости)
        if not price_str:
            price_str = "0"
        try:
            price = float(price_str)
        except ValueError:
            continue
        if price < 0 or price > MAX_PRICE:
            continue
        results.append({"barcode": barcode, "price": price})

    return results


def _parse_xlsx_content(content: bytes) -> list[dict]:
    """Парсит XLSX из bytes, возвращает list[{id, price}]."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="XLSX файл пуст")

    rows_iter = ws.iter_rows(values_only=True)

    # Ищем строку с заголовками (может быть не первая — шаблон имеет title row)
    id_col = None
    price_col = None
    for candidate_row in rows_iter:
        if not candidate_row:
            continue
        headers = [str(h).strip().lower() if h else "" for h in candidate_row]
        id_col = next((i for i, h in enumerate(headers) if h in ID_ALIASES), None)
        price_col = next((i for i, h in enumerate(headers) if h in PRICE_ALIASES), None)
        if id_col is not None and price_col is not None:
            break  # Нашли заголовки
        # Fallback: если два столбца и похоже на данные (первый = строка, последний = число)
        if len(headers) >= 2:
            try:
                float(str(candidate_row[-1]))
                # Это данные без заголовков — не пропускаем
                id_col = 0
                price_col = len(headers) - 1
                # Вернуть эту строку обратно нельзя, но она пойдёт как первая данная
                break
            except (ValueError, TypeError):
                pass  # Не число — пропускаем (title row)

    if id_col is None or price_col is None:
        raise HTTPException(
            status_code=400,
            detail="Не найдены колонки barcode/offer_id и purchase_price. Проверьте заголовки.",
        )

    results = []
    for row in rows_iter:
        if not row or len(row) <= max(id_col, price_col):
            continue
        barcode = str(row[id_col]).strip() if row[id_col] is not None else ""
        if not barcode:
            continue
        price_val = row[price_col]
        # Пустая ячейка или None = 0 (удаление себестоимости)
        if price_val is None or (isinstance(price_val, str) and price_val.strip() == ""):
            price_val = 0
        try:
            price = float(price_val)
        except (ValueError, TypeError):
            continue
        if price < 0 or price > MAX_PRICE:
            continue
        results.append({"barcode": barcode, "price": price})

    wb.close()
    return results


@router.post("/products/import-prices")
async def import_prices(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(get_current_user),
):
    """
    Импорт себестоимости из CSV или XLSX файла.
    Мэтчит строки по barcode, обновляет purchase_price.
    """
    # 0. Validate file extension
    allowed_extensions = (".csv", ".xlsx", ".xls")
    filename = file.filename or ""
    if not filename.lower().endswith(allowed_extensions):
        raise HTTPException(
            status_code=400,
            detail=f"Неподдерживаемый формат файла. Допустимые: {', '.join(allowed_extensions)}",
        )

    # 1. Читаем и валидируем размер
    content = await file.read()
    if len(content) > MAX_IMPORT_FILE_SIZE:
        raise HTTPException(status_code=400, detail="Файл превышает 5MB")
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Файл пуст")

    # 2. Определяем формат по magic bytes
    is_xlsx = content[:2] == b"PK"

    # 3. Парсим
    if is_xlsx:
        parsed_rows = _parse_xlsx_content(content)
    else:
        parsed_rows = _parse_csv_content(content)

    if not parsed_rows:
        raise HTTPException(status_code=400, detail="Не удалось распарсить ни одной строки с данными")

    if len(parsed_rows) > MAX_IMPORT_ROWS:
        raise HTTPException(status_code=400, detail=f"Превышен лимит строк ({MAX_IMPORT_ROWS})")

    # 4. Загружаем товары пользователя
    supabase = get_supabase_client()
    products_result = (
        supabase.table("mp_products")
        .select("id, barcode, purchase_price, product_group_id")
        .eq("user_id", current_user.id)
        .neq("barcode", "WB_ACCOUNT")
        .limit(2000)
        .execute()
    )

    by_barcode: dict[str, dict] = {}
    for p in products_result.data or []:
        by_barcode[p["barcode"]] = p

    # 5. Мэтчим и собираем batch updates
    import json

    skipped = 0
    not_found: list[str] = []
    errors: list[str] = []

    # Собираем обновления для batch RPC
    batch_updates: list[dict] = []
    group_updates: dict[str, float] = {}  # group_id → price

    for row in parsed_rows:
        barcode = row["barcode"]
        price = row["price"]

        product = by_barcode.get(barcode)
        if not product:
            not_found.append(barcode)
            continue

        if product["purchase_price"] == price:
            skipped += 1
            continue

        batch_updates.append({"id": product["id"], "purchase_price": price})

        gid = product.get("product_group_id")
        if gid:
            group_updates[gid] = price

    # Batch update via RPC
    updated = 0
    if batch_updates:
        try:
            result = supabase.rpc("batch_update_products", {
                "p_user_id": current_user.id,
                "p_updates": batch_updates,
            }).execute()
            updated = result.data if isinstance(result.data, int) else len(batch_updates)
        except Exception:
            # Fallback: поштучный UPDATE
            for upd in batch_updates:
                try:
                    supabase.table("mp_products").update({
                        "purchase_price": upd["purchase_price"],
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }).eq("id", upd["id"]).eq("user_id", current_user.id).execute()
                    updated += 1
                except Exception as e2:
                    errors.append(f"update {upd['id']}: {str(e2)}")

    # Update grouped products (linked pairs get same price)
    for gid, price in group_updates.items():
        try:
            supabase.table("mp_products").update({
                "purchase_price": price,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }).eq("product_group_id", gid).eq("user_id", current_user.id).execute()
        except Exception as e:
            logger.warning("import-prices group update failed for %s: %s", gid, e)

    return {
        "status": "success",
        "updated": updated,
        "total_rows": len(parsed_rows),
        "skipped": skipped,
        "not_found": not_found[:50],  # лимитируем вывод
        "errors": errors[:20],
    }


@router.post("/products/unlink/{group_id}")
async def unlink_products(
    group_id: str,
    current_user: CurrentUser = Depends(get_current_user),
):
    """
    Разорвать связь между товарами в группе.
    purchase_price остаётся как есть.
    """
    supabase = get_supabase_client()

    try:
        now = datetime.now(timezone.utc).isoformat()

        result = (
            supabase.table("mp_products")
            .update({"product_group_id": None, "updated_at": now})
            .eq("product_group_id", group_id)
            .eq("user_id", current_user.id)
            .execute()
        )

        unlinked_count = len(result.data) if result.data else 0

        if unlinked_count == 0:
            raise HTTPException(status_code=404, detail="No products found with this group_id")

        return {"status": "success", "unlinked_count": unlinked_count}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
